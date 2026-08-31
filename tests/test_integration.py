"""端到端集成测试"""
import os
import pytest
from openpyxl import load_workbook
from db import get_connection
from queries import get_available_months, get_salary_records, get_deduction_details, get_personnel_info
from models import MonthOption, SalaryRecord, PersonnelInfo
from templates_gen.normal_salary import generate_normal_salary
from templates_gen.labor_service import generate_labor_service
from templates_gen.annual_bonus import generate_annual_bonus
from templates_gen.personnel_info import generate_personnel_info
from templates_gen.validation import validate_salary_records


class TestQueryFunctions:
    """测试查询函数"""
    
    def test_get_available_months(self, conn):
        """测试月份查询"""
        months = get_available_months(conn)
        assert len(months) > 0
        assert isinstance(months[0], MonthOption)
        assert months[0].value > 202000  # 应该是合理的年月格式
    
    def test_get_salary_records(self, conn):
        """测试工资记录查询"""
        # 使用最近的月份
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        assert len(records) > 0
        assert isinstance(records[0], SalaryRecord)
        assert records[0].姓名 is not None
        assert records[0].职工号 is not None
    
    def test_get_salary_records_fields(self, conn):
        """测试工资记录字段完整性"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        rec = records[0]
        # 关键字段应该存在
        assert hasattr(rec, '应发工资')
        assert hasattr(rec, '养老个人')
        assert hasattr(rec, '医疗个人')
        assert hasattr(rec, '失业个人')
        assert hasattr(rec, '公积金个人')
        assert hasattr(rec, '实发工资')
        assert hasattr(rec, '个人所得税')
        assert hasattr(rec, '补缴及退款保险金额个人')
        assert hasattr(rec, '大病险个人')
        assert hasattr(rec, '采暖费')
        assert hasattr(rec, '独生子女费')
    
    def test_get_personnel_info(self, conn):
        """测试人员信息查询"""
        months = get_available_months(conn)
        latest_month = months[0].value
        personnel = get_personnel_info(conn, latest_month)
        assert len(personnel) > 0
        assert isinstance(personnel[0], PersonnelInfo)
        assert personnel[0].姓名 is not None


class TestTemplateGeneration:
    """测试模板生成"""
    
    def test_generate_normal_salary(self, conn, output_dir):
        """测试正常工资薪金所得模板生成"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        result = generate_normal_salary(records, f"测试工资{latest_month}", output_dir)
        assert result.file_path is not None
        assert os.path.exists(result.file_path)
        assert result.record_count == len(records)
        assert result.template_type == "正常工资薪金所得"
        # 验证 Excel 内容
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.max_row == len(records) + 1  # 表头 + 数据行
        assert ws.max_column == 29  # 29 列
    
    def test_generate_labor_service(self, conn, output_dir):
        """测试劳务报酬所得模板生成"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        result = generate_labor_service(records, f"测试劳务{latest_month}", output_dir)
        assert result.file_path is not None
        assert os.path.exists(result.file_path)
        assert result.record_count == len(records)
        assert result.template_type == "劳务报酬所得"
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.max_row == len(records) + 1
        assert ws.max_column == 14
    
    def test_generate_annual_bonus(self, conn, output_dir):
        """测试全年一次性奖金收入模板生成"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        result = generate_annual_bonus(records, f"测试奖金{latest_month}", output_dir)
        assert result.file_path is not None
        assert os.path.exists(result.file_path)
        assert result.template_type == "全年一次性奖金收入"
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.max_column == 11
    
    def test_generate_personnel_info(self, conn, output_dir):
        """测试人员信息采集导入模板生成"""
        months = get_available_months(conn)
        latest_month = months[0].value
        personnel = get_personnel_info(conn, latest_month)
        result = generate_personnel_info(personnel, f"测试人员{latest_month}", output_dir)
        assert result.file_path is not None
        assert os.path.exists(result.file_path)
        assert result.record_count == len(personnel)
        assert result.template_type == "人员信息采集导入模板"
        wb = load_workbook(result.file_path)
        ws = wb.active
        assert ws.max_column == 51


class TestValidation:
    """测试验证功能"""
    
    def test_validate_salary_records(self, conn):
        """测试工资记录验证"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        report = validate_salary_records(records)
        assert report.total_count == len(records)
        assert report.pass_count + report.fail_count == report.total_count
        assert report.pass_count >= 0
        assert report.fail_count >= 0
    
    def test_validation_report_structure(self, conn):
        """测试验证报告结构"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        report = validate_salary_records(records)
        # 如果有失败记录，检查详情结构
        if report.fail_count > 0:
            detail = report.details[0]
            assert '姓名' in detail
            assert '职工号' in detail
            assert '左' in detail
            assert '右' in detail
            assert '差值' in detail


class TestFlaskAPI:
    """测试 Flask API"""
    
    def test_months_api(self, app_client):
        """测试月份 API"""
        resp = app_client.get('/api/months')
        assert resp.status_code == 200
        data = resp.get_json()
        assert len(data) > 0
        assert 'value' in data[0]
        assert 'label' in data[0]
    
    def test_generate_api(self, app_client):
        """测试生成 API"""
        resp = app_client.post('/api/generate', 
            json={"month": 202607, "templates": ["normalSalary"]})
        assert resp.status_code == 200
        data = resp.get_json()
        assert 'files' in data
        assert len(data['files']) > 0
        assert 'download_url' in data['files'][0]
    
    def test_validate_api(self, app_client):
        """测试验证 API"""
        resp = app_client.get('/api/validate/202607')
        assert resp.status_code == 200
        data = resp.get_json()
        assert 'total_count' in data
        assert 'pass_count' in data
        assert 'fail_count' in data
    
    def test_download_api(self, app_client):
        """测试下载 API"""
        # 先生成文件
        resp = app_client.post('/api/generate', 
            json={"month": 202607, "templates": ["normalSalary"]})
        data = resp.get_json()
        filename = data['files'][0]['name']
        # 下载文件
        resp = app_client.get(f'/api/download/{filename}')
        assert resp.status_code == 200
        assert resp.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def test_error_handling(self, app_client):
        """测试错误处理"""
        resp = app_client.post('/api/generate', 
            json={"month": 999999, "templates": ["normalSalary"]})
        assert resp.status_code in (200, 400)
        
        resp = app_client.post('/api/generate', 
            json={"month": 202607, "templates": ["invalidTemplate"]})
        assert resp.status_code == 200


class TestEdgeCases:
    """测试边缘情况"""
    
    def test_empty_month(self, app_client):
        """测试空月份"""
        resp = app_client.post('/api/generate', 
            json={"month": None, "templates": ["normalSalary"]})
        assert resp.status_code == 400
    
    def test_no_templates(self, app_client):
        """测试未选择模板"""
        resp = app_client.post('/api/generate', 
            json={"month": 202607, "templates": []})
        assert resp.status_code == 400
    
    def test_missing_file_download(self, app_client):
        """测试下载不存在的文件"""
        resp = app_client.get('/api/download/nonexistent.xlsx')
        assert resp.status_code == 404


class TestModelCompleteness:
    """测试模型字段完整性，防止遗漏字段"""

    REQUIRED_SALARY_FIELDS = [
        '职工号', '姓名', '身份证', '工资所属年月', '结算单元', '当月批次', 'tc930_id',
        '应发工资', '实发工资', '个人所得税', '工资总额',
        '独生子女费', '采暖费', '奖金',
        '养老个人', '医疗个人', '失业个人', '公积金个人',
        '补缴及退款保险金额个人', '大病险个人', '补发3', '个人交纳现金',
    ]

    def test_salary_record_has_all_fields(self):
        """SalaryRecord 必须包含所有必要字段"""
        rec = SalaryRecord()
        for field in self.REQUIRED_SALARY_FIELDS:
            assert hasattr(rec, field), f"SalaryRecord 缺少字段: {field}"

    def test_salary_record_tc930_id(self):
        """SalaryRecord.tc930_id 必须可正常赋值和读取"""
        rec = SalaryRecord(tc930_id=12345)
        assert rec.tc930_id == 12345


class TestGenerateEndToEnd:
    """测试生成端到端流程，覆盖所有新增sheet"""

    def test_generate_all_sheets(self, conn, output_dir):
        """generate_normal_salary 必须生成: 正常工资薪金收入、验证报告"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        result = generate_normal_salary(records, f"测试{latest_month}", output_dir)
        wb = load_workbook(result.file_path)
        assert "正常工资薪金收入" in wb.sheetnames
        assert "验证报告" in wb.sheetnames

    def test_generate_with_tc930_in_validation(self, conn, output_dir):
        """验证报告 sheet 必须包含 ATC930 列"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        result = generate_normal_salary(records, f"测试{latest_month}", output_dir)
        wb = load_workbook(result.file_path)
        vs = wb["验证报告"]
        headers = [vs.cell(row=1, column=c).value for c in range(1, vs.max_column + 1)]
        assert "ATC930" in headers

    def test_records_have_tc930_id(self, conn):
        """查询到的每条记录 tc930_id 必须非零"""
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        for rec in records:
            assert rec.tc930_id != 0, f"记录 {rec.姓名} tc930_id 为0"

    def test_generate_api_returns_abnormal_count(self, app_client):
        """generate API 必须返回 abnormal_count 和 tc93_total_count"""
        resp = app_client.post('/api/generate',
            json={"month": 202607, "templates": ["normalSalary"]})
        assert resp.status_code == 200
        data = resp.get_json()
        assert 'abnormal_count' in data
        assert 'tc93_total_count' in data
        assert data['tc93_total_count'] >= data['abnormal_count']

    def test_generate_api_default_merge_by_pay_month(self, app_client, conn):
        """未传 merge_by_pay_month 时默认按人+发放月份合并: 导出行数=唯一人数"""
        from queries import get_salary_records
        raw = get_salary_records(conn, 202607)
        unique_persons = len({r.职工号 for r in raw})
        resp = app_client.post('/api/generate',
            json={"month": 202607, "templates": ["normalSalary"]})
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['files'][0]['count'] == unique_persons

    def test_generate_with_combos(self, conn, output_dir):
        """带 confirmed_combos 时 TC93总表和异常记录表必须按组合过滤"""
        from queries import get_tc93_all_fields, get_abnormal_records
        from templates_gen.normal_salary import generate_normal_salary
        months = get_available_months(conn)
        latest_month = months[0].value
        records = get_salary_records(conn, latest_month)
        if not records:
            pytest.skip("无工资记录")
        first = records[0]
        combos = [{"unit": first.结算单元, "salary_month": first.工资所属年月, "seq": first.当月批次}]
        combo_set = {(c["unit"], c["salary_month"], c["seq"]) for c in combos}
        tc93_all = [r for r in get_tc93_all_fields(conn, latest_month)
                     if (r.get("ATB930"), r.get("ATC931"), r.get("ATC937")) in combo_set]
        abnormal = [r for r in get_abnormal_records(conn, latest_month)
                     if (r.get("ATB930"), r.get("ATC931"), r.get("ATC937")) in combo_set]
        result = generate_normal_salary(records, f"测试{latest_month}", output_dir,
                                         tc93_all=tc93_all, abnormal=abnormal,
                                         abnormal_reasons={}, combos=combos)
        wb = load_workbook(result.file_path)
        if "TC93总表" in wb.sheetnames:
            ws = wb["TC93总表"]
            assert ws.max_row <= len(tc93_all) + 2


class TestX3Income:
    """个人交纳现金(ATC93X3)计入本期收入, 与个税端回盘口径一致"""

    def _mk(self, total=0, be=0, x3=0, social=0, paid=0, tax=0):
        from decimal import Decimal
        rec = SalaryRecord()
        rec.工资总额 = Decimal(str(total))
        rec.补缴及退款保险金额个人 = Decimal(str(be))
        rec.个人交纳现金 = Decimal(str(x3))
        rec.养老个人 = Decimal(str(social))
        rec.实发工资 = Decimal(str(paid))
        rec.个人所得税 = Decimal(str(tax))
        return rec

    def test_x3_added_to_income(self):
        """宋红玥案例: 应发0 交纳现金608.5 -> 本期收入608.5"""
        from decimal import Decimal
        from templates_gen.formulas import calc_本期收入
        rec = self._mk(total=0, x3=608.5)
        assert calc_本期收入(rec) == Decimal("608.5")

    def test_x3_balances_validation(self):
        """宋霜案例: 总额359.68 X3 608.5 五险608.5 实发359.68 -> 左=右"""
        from decimal import Decimal
        from templates_gen.formulas import calc_本期收入, calc_免税
        rec = self._mk(total=359.68, x3=608.5, social=608.5, paid=359.68)
        income = calc_本期收入(rec)
        assert income == Decimal("968.18")
        left = income - rec.养老个人
        right = rec.实发工资 + rec.个人所得税 - calc_免税(rec)
        assert abs(left - right) < Decimal("0.01")

    def test_merge_sums_x3(self):
        """合并时个人交纳现金必须合计"""
        from decimal import Decimal
        from app import merge_records_by_person
        recs = [self._mk(total=0, x3=608.5), self._mk(total=100, x3=200)]
        recs[0].职工号 = recs[1].职工号 = "1"
        merged = merge_records_by_person(recs)
        assert len(merged) == 1
        assert merged[0].个人交纳现金 == Decimal("808.5")

    def test_x3_and_e_match_return_disk(self):
        """马静案例: 总额0 BE1825.5 X3=0 E=-2434 -> 收入608.5(与回盘一致)"""
        from decimal import Decimal
        from templates_gen.formulas import calc_本期收入
        rec = self._mk(total=0, be=1825.5, social=608.5)
        rec.个人欠款 = Decimal("-2434")
        assert calc_本期收入(rec) == Decimal("608.5")

    def test_x3_e_validation_equivalent(self):
        """E 从左式移入收入后, 左=右 结果与原公式代数等价"""
        from decimal import Decimal
        from templates_gen.formulas import calc_本期收入, calc_免税
        rec = self._mk(total=0, be=1825.5, social=608.5)
        rec.个人欠款 = Decimal("-2434")
        income = calc_本期收入(rec)
        # 新左式(不含E)
        left_new = income - rec.养老个人
        # 旧左式(收入不含E, 左式减E)
        income_old = rec.工资总额 - rec.补缴及退款保险金额个人
        left_old = income_old - rec.养老个人 - rec.个人欠款
        assert left_new == left_old
        # 零工资补缴行: 左=0=右
        right = rec.实发工资 + rec.个人所得税 - calc_免税(rec)
        assert abs(left_new - right) < Decimal("0.01")


class TestMergeByPayMonth:
    """按人+发放月份合并（同一发放月份每人一行，跨所属月份收入/五险一金/个税合计）"""

    def _mk(self, pid, name, month, batch, total=0, tax=0, social=0, paid=0,
            cert="130225198812115529", unit=37339):
        from decimal import Decimal
        rec = SalaryRecord(职工号=pid, 姓名=name, 身份证=cert, 工资所属年月=month,
                           结算单元=unit, 当月批次=batch, tc930_id=month * 100 + int(batch))
        rec.工资总额 = Decimal(str(total))
        rec.个人所得税 = Decimal(str(tax))
        rec.养老个人 = Decimal(str(social))
        rec.实发工资 = Decimal(str(paid))
        return rec

    def _wang4(self):
        # 与真实数据一致: 王建秋 4 笔跨所属月份记录
        return [
            self._mk("4370", "王建秋", 202604, 1),
            self._mk("4370", "王建秋", 202605, 3),
            self._mk("4370", "王建秋", 202606, 4, total=9210.6, tax=432.8,
                     social=2046.59, paid=6731.21),
            self._mk("4370", "王建秋", 202607, "2", total=61978.65, tax=13733.01,
                     social=2046.59, paid=46199.05),
        ]

    def test_merge_by_pay_month_groups_person_only(self):
        """by_pay_month=True 时 4 笔跨月记录合并为 1 笔，金额全部合计"""
        from decimal import Decimal
        from app import merge_records_by_person
        merged = merge_records_by_person(self._wang4(), by_pay_month=True)
        assert len(merged) == 1
        m = merged[0]
        assert m.工资总额 == Decimal("71189.25")
        assert m.个人所得税 == Decimal("14165.81")
        assert m.养老个人 == Decimal("4093.18")
        assert m.实发工资 == Decimal("52930.26")
        # 基准取流水号最大(最新经办)记录: 202607 批2
        assert m.工资所属年月 == 202607
        assert m.当月批次 == "2"

    def test_merge_default_keeps_month_granularity(self):
        """默认(by_pay_month=False)仍按人+所属月份合并，跨月不合并"""
        from app import merge_records_by_person
        merged = merge_records_by_person(self._wang4())
        assert len(merged) == 4

    def test_merge_pay_month_export_one_row_per_person(self, output_dir):
        """merge_mode=pay_month 时 正常工资薪金收入 每人一行，合并明细可追溯"""
        from decimal import Decimal
        from app import merge_records_by_person
        raw = self._wang4()
        merged = merge_records_by_person(raw, by_pay_month=True)
        result = generate_normal_salary(
            merged, "测试按人合并", output_dir, raw_records=raw, merge_mode="pay_month",
            combos=[{"unit": 37339, "salary_month": 202607, "seq": "2",
                     "unit_name": "吉林大学第二医院B"}])
        wb = load_workbook(result.file_path)
        ws = wb["正常工资薪金收入"]
        assert ws.max_row == 2  # 表头 + 1人1行
        assert ws.cell(row=2, column=5).value == 71189.25  # 本期收入合计
        assert ws.cell(row=2, column=7).value == 4093.18  # 养老合计
        assert "吉林大学第二医院B-202607-2" in str(ws.cell(row=2, column=29).value)
        md = wb["合并明细"]
        assert md.max_row == 2
        assert md.cell(row=2, column=4).value == 4  # 原始条数
        assert md.cell(row=2, column=3).value == "202604;202605;202606;202607"
        assert md.cell(row=2, column=9).value == 4093.18  # 合并五险
        assert md.cell(row=2, column=10).value == 14165.81  # 合并个税
