"""tax_return.py - 税务局个税端回盘数据解析、存储与报税状态比对"""
import os
import sqlite3
from datetime import datetime
from decimal import Decimal

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tax_return.db")

HEADER_MAP = {
    "姓名": "name",
    "证件号码": "cert_no",
    "税款所属期起": "period_start",
    "本期收入": "income",
    "本期免税收入": "tax_free",
    "本期基本养老保险费": "pension",
    "本期基本医疗保险费": "medical",
    "本期失业保险费": "unemployment",
    "本期住房公积金": "housing",
    "累计应扣缴税额": "tax_accum",
    "已缴税额": "tax_paid",
    "应补(退)税额": "tax_due",
    "备注": "remark",
}


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS tax_return (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cert_no TEXT NOT NULL,
            name TEXT NOT NULL,
            month INTEGER NOT NULL,
            income REAL DEFAULT 0,
            tax_free REAL DEFAULT 0,
            insurance REAL DEFAULT 0,
            tax_accum REAL DEFAULT 0,
            tax_paid REAL DEFAULT 0,
            tax_due REAL DEFAULT 0,
            remark TEXT DEFAULT '',
            import_time TEXT,
            UNIQUE(cert_no, month)
        )
    """)
    conn.commit()
    conn.close()


def _as_decimal(v):
    try:
        return float(str(v or 0))
    except (ValueError, TypeError):
        return 0.0


def parse_return_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        rows = _read_csv(filepath)
    elif ext in (".xls", ".xlsx"):
        rows = _read_excel(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")
    return _normalize(rows)


def _read_csv(filepath):
    import csv
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        return list(reader)


def _read_excel(filepath):
    if filepath.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _normalize(rows):
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    idx = {}
    for i, h in enumerate(header):
        if h in HEADER_MAP:
            idx[HEADER_MAP[h]] = i
    if "cert_no" not in idx or "name" not in idx:
        raise ValueError("回盘文件缺少 姓名/证件号码 列")

    records = []
    for row in rows[1:]:
        cert = str(row[idx["cert_no"]] or "").strip()
        if not cert:
            continue
        period = str(row[idx.get("period_start", 0)] or "")
        month = int(period.replace("-", "")[:6]) if period else 0
        insurance = sum(_as_decimal(row[idx[k]]) for k in ("pension", "medical", "unemployment", "housing") if k in idx)
        records.append({
            "cert_no": cert,
            "name": str(row[idx["name"]] or "").strip(),
            "month": month,
            "income": _as_decimal(row[idx["income"]]) if "income" in idx else 0,
            "tax_free": _as_decimal(row[idx["tax_free"]]) if "tax_free" in idx else 0,
            "insurance": insurance,
            "tax_accum": _as_decimal(row[idx["tax_accum"]]) if "tax_accum" in idx else 0,
            "tax_paid": _as_decimal(row[idx["tax_paid"]]) if "tax_paid" in idx else 0,
            "tax_due": _as_decimal(row[idx["tax_due"]]) if "tax_due" in idx else 0,
            "remark": str(row[idx["remark"]] or "").strip() if "remark" in idx else "",
        })
    return records


def import_records(records):
    conn = get_db()
    now = datetime.now().isoformat(timespec="seconds")
    for rec in records:
        if not rec["month"]:
            continue
        conn.execute("""
            INSERT OR REPLACE INTO tax_return
            (cert_no, name, month, income, tax_free, insurance, tax_accum, tax_paid, tax_due, remark, import_time)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (rec["cert_no"], rec["name"], rec["month"], rec["income"], rec["tax_free"],
              rec["insurance"], rec["tax_accum"], rec["tax_paid"], rec["tax_due"], rec["remark"], now))
    conn.commit()
    conn.close()


def get_returns(month):
    conn = get_db()
    rows = conn.execute("SELECT * FROM tax_return WHERE month=?", (month,)).fetchall()
    conn.close()
    return {r["cert_no"]: dict(r) for r in rows}


def get_system_aggregate(conn, pay_month):
    """按发放月份(税款所属期)聚合系统预算数据。

    pay_month 对应 TC8M.ATC8G7(发放月份/税款所属期)，先取该发放月份下的
    全部已发放组合(结算单元+所属月份+批次)，再按组合的所属月份分别查 TC93 聚合。
    """
    from queries import get_salary_records, search_tc8m
    from templates_gen.formulas import calc_本期收入, calc_免税
    tc8m_combos = search_tc8m(conn, pay_month=pay_month, status=2)
    combo_set = {(c["unit"], c["salary_month"], c["seq"]) for c in tc8m_combos}
    salary_months = {c["salary_month"] for c in tc8m_combos}
    agg = {}
    combos = {}
    for sm in salary_months:
        for rec in get_salary_records(conn, sm):
            if (rec.结算单元, rec.工资所属年月, rec.当月批次) not in combo_set:
                continue
            key = rec.身份证 or rec.职工号
            if not key:
                continue
            combo = (rec.结算单元, rec.工资所属年月, rec.当月批次)
            combos.setdefault(combo, {"persons": set(), "count": 0})
            combos[combo]["persons"].add(key)
            combos[combo]["count"] += 1
            cur = agg.get(key)
            income = calc_本期收入(rec)
            tax_exempt = calc_免税(rec)
            if cur is None:
                agg[key] = {
                    "name": rec.姓名,
                    "income": income,
                    "tax_free": tax_exempt,
                    "insurance": rec.养老个人 + rec.医疗个人 + rec.失业个人 + rec.公积金个人,
                    "tax": rec.个人所得税,
                    "unit": rec.结算单元,
                    "unit_keys": {combo},
                    "entries": [{"unit": rec.结算单元, "month": rec.工资所属年月,
                                 "seq": rec.当月批次, "income": float(income), "tax": float(rec.个人所得税)}],
                    "month": rec.工资所属年月,
                    "seq": rec.当月批次,
                }
            else:
                cur["income"] += income
                cur["tax_free"] += tax_exempt
                cur["insurance"] += rec.养老个人 + rec.医疗个人 + rec.失业个人 + rec.公积金个人
                cur["tax"] += rec.个人所得税
                cur["unit_keys"].add(combo)
                cur["entries"].append({"unit": rec.结算单元, "month": rec.工资所属年月,
                                       "seq": rec.当月批次, "income": float(income), "tax": float(rec.个人所得税)})
    return agg, combos


def compare_month(conn, month):
    """比对系统预算 vs 回盘，返回按人明细与按组合统计。"""
    system, combos = get_system_aggregate(conn, month)
    returns = get_returns(month)
    all_unit_keys = {k for s in system.values() for k in s.get("unit_keys", {(s.get("unit", 0), month, "1")})}
    unit_names = _get_unit_names(conn, list(all_unit_keys))
    details = []
    for cert, s in system.items():
        keys = s.get("unit_keys", {(s.get("unit", 0), month, "1")})
        by_unit = {}
        for k in sorted(keys):
            by_unit.setdefault(k[0], []).append((k[1], k[2], unit_names.get(k, "") or str(k[0])))
        unit_names_list = []
        periods_list = []
        for u, items in by_unit.items():
            unit_names_list.append(items[0][2])
            periods_list.append("、".join(f"{m}-批{s}" for m, s, _ in sorted(items, key=lambda x: (x[0], str(x[1])))))
        uname = "、".join(unit_names_list)
        uperiods = " | ".join(periods_list)
        entries = []
        for e in s.get("entries", []):
            ekey = (e["unit"], e["month"], e["seq"])
            entries.append({
                "unit_name": unit_names.get(ekey, "") or str(e["unit"]),
                "month": e["month"], "seq": e["seq"],
                "income": round(e["income"], 2), "tax": round(e["tax"], 2),
            })
        entries.sort(key=lambda x: (x["month"], str(x["seq"])))
        r = returns.get(cert)
        if r is None:
            details.append({
                "cert_no": cert, "name": s["name"], "month": month, "unit_name": uname, "unit_periods": uperiods,
                "entries": entries,
                "sys_income": round(float(s["income"]), 2), "sys_tax": round(float(s["tax"]), 2),
                "ret_tax": None, "diff": None, "status": "未报送",
            })
        else:
            diff = round(abs(float(s["tax"]) - float(r["tax_due"])), 2)
            income_diff = round(abs(float(s["income"]) - float(r["income"])), 2)
            details.append({
                "cert_no": cert, "name": s["name"], "month": month, "unit_name": uname, "unit_periods": uperiods,
                "entries": entries,
                "sys_income": round(float(s["income"]), 2), "ret_income": r["income"], "income_diff": income_diff,
                "sys_tax": round(float(s["tax"]), 2), "ret_tax": r["tax_due"], "diff": diff,
                "status": "已报送" if diff <= 0.01 else "有差异",
            })
    for cert, r in returns.items():
        if cert not in system:
            details.append({
                "cert_no": cert, "name": r["name"], "month": month, "unit_name": r.get("remark", ""), "unit_periods": "",
                "sys_income": None, "sys_tax": None,
                "ret_income": r["income"], "ret_tax": r["tax_due"], "diff": None,
                "status": "系统无记录",
            })

    combo_stats = []
    unit_names = _get_unit_names(conn, list(combos.keys()))
    for (unit, sm, seq), info in sorted(combos.items()):
        reported = sum(1 for p in info["persons"] if p in returns)
        combo_stats.append({
            "unit": unit,
            "unit_name": unit_names.get((unit, sm, seq), ""),
            "month": sm,
            "seq": seq,
            "count": info["count"],
            "reported": reported,
            "unreported": info["count"] - reported,
        })
    return details, combo_stats


def _get_unit_names(conn, combos):
    """从 TC8M 查询组合对应的结算单元名称(按银行拆分多条时取MAX去重)。"""
    if not combos:
        return {}
    names = {}
    for start in range(0, len(combos), 400):
        chunk = combos[start:start + 400]
        placeholders = ", ".join(f"(:u{i}, :m{i}, :s{i})" for i in range(len(chunk)))
        binds = {}
        for i, (u, m, s) in enumerate(chunk):
            binds[f"u{i}"] = u
            binds[f"m{i}"] = m
            binds[f"s{i}"] = s
        sql = f"""
            SELECT ATB930, ATC931, ATC937, MAX(ATB931)
            FROM TC8M
            WHERE (ATB930, ATC931, ATC937) IN ({placeholders})
            GROUP BY ATB930, ATC931, ATC937
        """
        with conn.cursor() as cursor:
            cursor.execute(sql, binds)
            for row in cursor.fetchall():
                names[(int(row[0] or 0), int(row[1] or 0), str(row[2] or ""))] = str(row[3] or "")
    return names


def summarize(conn, month):
    details, combo_stats = compare_month(conn, month)
    total = len(details)
    reported = sum(1 for d in details if d["status"] == "已报送")
    diff = sum(1 for d in details if d["status"] == "有差异")
    unreported = sum(1 for d in details if d["status"] == "未报送")
    no_sys = sum(1 for d in details if d["status"] == "系统无记录")
    return {
        "month": month,
        "total": total,
        "reported": reported,
        "diff": diff,
        "unreported": unreported,
        "no_sys": no_sys,
        "combos": combo_stats,
        "details": details,
    }