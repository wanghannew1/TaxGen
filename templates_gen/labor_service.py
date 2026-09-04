"""劳务报酬所得模板生成器 - 精确复制 demo TemplateFiller.cs 算法"""
from datetime import datetime
from typing import List
import os
from openpyxl import Workbook
from models import SalaryRecord, GenerateResult
from templates_gen.explanation import add_explanation_sheet


def extract_remark(title: str) -> str:
    """从标题提取备注字段"""
    if not title:
        return ""
    t = title.replace("东北师范大学人事处", "").replace("劳务派遣人员工资发放表", "").strip()
    t = t.replace("年", "").replace("月", "").replace("系统", "").strip()
    digits = ''.join(c for c in t if c.isdigit())
    if len(digits) >= 4:
        t = t.replace(digits, "").strip()
    return t if t else title


def generate_labor_service(records: List[SalaryRecord], title: str, output_dir: str) -> GenerateResult:
    """生成劳务报酬所得 Excel 模板"""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(output_dir, f"劳务报酬所得_{timestamp}.xlsx")
    remark = extract_remark(title)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "劳务报酬"
    
    # 14列标题 - 必须与 demo 完全一致
    headers = [
        "工号", "*姓名", "*证件类型", "*证件号码", "*所得项目", "*收入",
        "免税收入", "商业健康保险", "税延养老保险", "其他",
        "允许扣除的税费", "减免税额", "协定减免", "备注"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # 写入数据
    for idx, rec in enumerate(records, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=rec.职工号)
        ws.cell(row=row, column=2, value=rec.姓名)
        ws.cell(row=row, column=3, value="居民身份证")
        ws.cell(row=row, column=4, value=rec.身份证)
        ws.cell(row=row, column=5, value="劳务报酬")
        ws.cell(row=row, column=6, value=rec.应发工资)
        ws.cell(row=row, column=14, value=remark)
    
    add_explanation_sheet(wb, [
        ("劳务报酬", [
            "14 列个税劳务报酬申报模板，一行为一人，无校验。",
            "*所得项目 恒为「劳务报酬」；收入取应发工资（原始值，不做扣减）。",
            "*证件类型 恒为「居民身份证」。备注从标题中提取（去除机构名/年月/数字后剩余文本）。",
        ]),
    ])
    
    wb.save(output_path)
    
    return GenerateResult(
        file_path=output_path,
        template_type="劳务报酬所得",
        record_count=len(records),
        validation_pass=0,
        validation_fail=0
    )
