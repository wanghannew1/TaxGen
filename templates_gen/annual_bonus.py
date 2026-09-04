"""全年一次性奖金收入模板生成器 - 精确复制 demo TemplateFiller.cs 算法"""
from datetime import datetime
from typing import List
import os
from openpyxl import Workbook
from models import SalaryRecord, GenerateResult
from templates_gen.explanation import add_explanation_sheet


def extract_remark(title: str) -> str:
    """从标题提取备注字段"""
    if not title:
        return ""
    t = title.replace("东北师范大学人事处", "").replace("劳务派遣人员工资发放表", "").strip()
    t = t.replace("年", "").replace("月", "").replace("系统", "").strip()
    digits = ''.join(c for c in t if c.isdigit())
    if len(digits) >= 4:
        t = t.replace(digits, "").strip()
    return t if t else title


def generate_annual_bonus(records: List[SalaryRecord], title: str, output_dir: str) -> GenerateResult:
    """生成全年一次性奖金收入 Excel 模板 - 只包含有奖金的记录"""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(output_dir, f"全年一次性奖金收入_{timestamp}.xlsx")
    remark = extract_remark(title)
    
    # 只包含奖金 > 0 的记录
    bonus_records = [r for r in records if r.奖金 > 0]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "全年一次性奖金"
    
    # 11列标题 - 必须与 demo 完全一致
    headers = [
        "工号", "*姓名", "*证件类型", "*证件号码", "*全年一次性奖金额",
        "免税收入", "其他", "准予扣除的捐赠额", "减免税额", "协定减免", "备注"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # 写入数据
    for idx, rec in enumerate(bonus_records, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=rec.职工号)
        ws.cell(row=row, column=2, value=rec.姓名)
        ws.cell(row=row, column=3, value="居民身份证")
        ws.cell(row=row, column=4, value=rec.身份证)
        ws.cell(row=row, column=5, value=rec.奖金)
        ws.cell(row=row, column=11, value=remark)
    
    add_explanation_sheet(wb, [
        ("全年一次性奖金", [
            "11 列个税全年一次性奖金申报模板，一行为一人，无校验。",
            "仅包含奖金数额 ＞ 0 的记录。",
            "*全年一次性奖金额 取奖金数额；*证件类型 恒为「居民身份证」。",
            "备注从标题中提取（去除机构名/年月/数字后剩余文本）。",
        ]),
    ])
    
    wb.save(output_path)
    
    return GenerateResult(
        file_path=output_path,
        template_type="全年一次性奖金收入",
        record_count=len(bonus_records),
        validation_pass=0,
        validation_fail=0
    )