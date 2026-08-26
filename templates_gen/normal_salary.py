"""正常工资薪金所得模板生成器 - 精确复制 demo TemplateFiller.cs 算法"""
from datetime import datetime
from typing import List, Optional
from collections import Counter
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from models import SalaryRecord, GenerateResult
from templates_gen.formulas import calc_本期收入, calc_免税


def extract_remark(title: str) -> str:
    """从标题提取备注字段"""
    if not title:
        return ""
    t = title.replace("东北师范大学人事处", "").replace("劳务派遣人员工资发放表", "").strip()
    t = t.replace("年", "").replace("月", "").replace("系统", "").strip()
    digits = ''.join(c for c in t if c.isdigit())
    if len(digits) >= 4:
        t = t.replace(digits, "").strip()
    return t if t else title


def generate_normal_salary(records: List[SalaryRecord], title: str, output_dir: str,
                           tc93_all: Optional[List[dict]] = None,
                           abnormal: Optional[List[dict]] = None,
                           abnormal_reasons: Optional[dict] = None,
                           combos: Optional[List[dict]] = None,
                           tc93_comments: Optional[dict] = None,
                           raw_records: Optional[List[SalaryRecord]] = None) -> GenerateResult:
    """生成正常工资薪金所得 Excel 模板

    新增 tc93_all: TC93总表(全字段), abnormal: 异常记录, abnormal_reasons: 过滤原因
    """
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(output_dir, f"正常工资薪金所得_{title}_{timestamp}.xlsx")

    combo_map = {}
    if combos:
        for c in combos:
            key = (int(c.get("unit", 0) or 0), int(c.get("salary_month", 0) or 0), str(c.get("seq", "") or ""))
            combo_map[key] = f"{c.get('unit_name', '')}-{c.get('salary_month', '')}-{c.get('seq', '')}"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "正常工资薪金收入"
    
    # 29列标题 - 必须与 demo 完全一致
    headers = [
        "工号", "*姓名", "*证件类型", "*证件号码", "本期收入", "本期免税收入",
        "基本养老保险费", "基本医疗保险费", "失业保险费", "住房公积金",
        "累计子女教育", "累计继续教育", "累计住房贷款利息", "累计住房租金",
        "累计赡养老人", "累计3岁以下婴幼儿照护", "累计个人养老金", "企业(职业)年金",
        "商业健康保险", "税延养老保险", "公务交通费用", "通讯费用", "律师办案费用",
        "西藏附加减除费用", "其他", "准予扣除的捐赠额", "减免税额", "协定减免", "备注"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # 写入数据
    validations = []
    for idx, rec in enumerate(records, 1):
        row = idx + 1
        income = calc_本期收入(rec)
        tax_exempt = calc_免税(rec)

        ws.cell(row=row, column=1, value=rec.职工号)
        ws.cell(row=row, column=2, value=rec.姓名)
        ws.cell(row=row, column=3, value="居民身份证")
        ws.cell(row=row, column=4, value=rec.身份证)
        ws.cell(row=row, column=5, value=income)
        ws.cell(row=row, column=7, value=rec.养老个人)
        ws.cell(row=row, column=8, value=rec.医疗个人)
        ws.cell(row=row, column=9, value=rec.失业个人)
        ws.cell(row=row, column=10, value=rec.公积金个人)
        ws.cell(row=row, column=18, value=0)  # 企业(职业)年金 = 0
        rec_remark = combo_map.get((rec.结算单元, rec.工资所属年月, rec.当月批次), title)
        ws.cell(row=row, column=29, value=rec_remark)
        
        left = income - rec.养老个人 - rec.失业个人 - rec.医疗个人 - rec.公积金个人 \
               - rec.个人其他调整 - rec.个人欠款 - rec.扣款大病险 - rec.意外险个人
        right = rec.实发工资 + rec.税后工会会费 + rec.个人代理费 + rec.个人所得税 - tax_exempt
        diff = abs(left - right)
        passed = diff < 0.01
        validations.append({
            "tc930": rec.tc930_id, "姓名": rec.姓名,
            "unit_name": combo_map.get((rec.结算单元, rec.工资所属年月, rec.当月批次), title),
            "salary_month": rec.工资所属年月, "seq": rec.当月批次,
            "工资总额": rec.工资总额, "本次免税": rec.补发3, "大病险个人": rec.大病险个人,
            "补缴退款差额": rec.补缴及退款保险金额个人, "本期收入": income,
            "养老": rec.养老个人, "失业": rec.失业个人, "医疗": rec.医疗个人, "公积金": rec.公积金个人,
            "其他调整": rec.个人其他调整, "个人欠款": rec.个人欠款, "扣款大病险": rec.扣款大病险, "意外险": rec.意外险个人,
            "左": left,
            "实发": rec.实发工资, "工会会费": rec.税后工会会费, "代理费": rec.个人代理费,
            "个税": rec.个人所得税, "免税": tax_exempt,
            "右": right, "差值": diff, "通过": passed
        })
    
    vs = wb.create_sheet("验证报告")
    vs_headers = [
        "ATC930", "姓名", "结算单元名称", "所属月份", "批次",
        "本次工资总额(ATC93AA)", "本次免税(ATC936)", "大病险（个人承担）(ATC93BD)", "补缴及退款保险差额（个人）(ATC93BE)", "本期收入",
        "当月养老个人缴(BAA001)", "当月失业个人缴(BAA003)", "当月医疗个人缴(BAA002)", "个人公积金月缴存额(CAA002)",
        "个人其他调整(ATC93AG)", "个人欠款(ATC93E)", "扣款-大病险(ATC93Y2)", "意外险个人(ATC93BH)", "左",
        "本次实发金额(ATC93C)", "税后扣除工会会费(ATC93Z2)", "个人承担代理费(BAA300)", "本次个人所得税(ATC93D)", "免税(ATC936+ATC93BD)",
        "右", "差值", "状态"
    ]
    for col, h in enumerate(vs_headers, 1):
        vs.cell(row=1, column=col, value=h)
    for idx, v in enumerate(validations, 1):
        vals = [
            v["tc930"], v["姓名"], v["unit_name"], v["salary_month"], v["seq"],
            v["工资总额"], v["本次免税"], v["大病险个人"], v["补缴退款差额"], v["本期收入"],
            v["养老"], v["失业"], v["医疗"], v["公积金"],
            v["其他调整"], v["个人欠款"], v["扣款大病险"], v["意外险"], v["左"],
            v["实发"], v["工会会费"], v["代理费"], v["个税"], v["免税"],
            v["右"], v["差值"], "通过" if v["通过"] else "失败"
        ]
        for col, val in enumerate(vals, 1):
            vs.cell(row=idx+1, column=col, value=val)
    
    if tc93_all:
        generate_tc93_full_sheet(wb, tc93_all, tc93_comments)
    if abnormal:
        generate_abnormal_sheet(wb, abnormal, abnormal_reasons or {})
    if combos:
        generate_combo_list_sheet(wb, combos)
    if raw_records:
        generate_raw_detail_sheet(wb, raw_records, combo_map, title)
        generate_merge_detail_sheet(wb, raw_records, records)
    generate_formula_explanation_sheet(wb, records)
    
    wb.save(output_path)
    
    pass_count = sum(1 for v in validations if v["通过"])
    fail_count = len(validations) - pass_count
    
    return GenerateResult(
        file_path=output_path,
        template_type="正常工资薪金所得",
        record_count=len(records),
        validation_pass=pass_count,
        validation_fail=fail_count
    )


def generate_tc93_full_sheet(wb: Workbook, tc93_all: List[dict], comments: Optional[dict] = None):
    """TC93总表 sheet：第1行字段注释，第2行字段名，数据从第3行。按身份证排序，身份证右侧加重复计数列。"""
    if not tc93_all:
        return
    ws = wb.create_sheet("TC93总表")
    cols = list(tc93_all[0].keys())
    id_counts = Counter(rec.get("身份证", "") for rec in tc93_all)
    if "身份证" in cols:
        idx = cols.index("身份证")
        cols.insert(idx + 1, "重复次数")
    for col_idx, col_name in enumerate(cols, 1):
        if col_name == "重复次数":
            ws.cell(row=1, column=col_idx, value="该身份证号在本sheet中出现的行数")
        else:
            comment = (comments or {}).get(col_name, "")
            ws.cell(row=1, column=col_idx, value=comment if comment else col_name)
        ws.cell(row=2, column=col_idx, value=col_name)
    for row_idx, rec in enumerate(tc93_all, 3):
        for col_idx, col_name in enumerate(cols, 1):
            if col_name == "重复次数":
                val = id_counts.get(rec.get("身份证", ""), 0)
            else:
                val = rec.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, str) and val.startswith("="):
                cell.data_type = "s"
    for col_idx, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(25, len(str(col_name)) * 1.5))


def generate_abnormal_sheet(wb: Workbook, abnormal: List[dict], reasons: dict):
    """异常记录sheet，列出被状态过滤的条目及原因。"""
    if not abnormal:
        return
    ws = wb.create_sheet("异常记录(已过滤)")
    base_cols = ["AAC001", "AAC003", "ATC931", "ATC937", "ATC930", "ATB930", "ATC93AA", "ATC93C", "ATC93D", "ATC93G", "ATC93N", "ATC93U", "ATC93V", "ATC93W", "ATC93AE"]
    headers = ["职工号", "姓名", "所属年月", "批次", "流水号", "结算单元", "工资总额", "实发金额", "个税", "结算状态", "上月工资", "可发情况", "费用状态", "个人欠费", "偿还", "过滤原因"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, rec in enumerate(abnormal, 2):
        tc930 = rec.get("ATC930")
        for col_idx, col_name in enumerate(base_cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(col_name))
        ws.cell(row=row_idx, column=len(base_cols) + 1, value=reasons.get(tc930, "状态异常"))
    # 自动列宽
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(25, len(str(h)) * 1.5))


def generate_combo_list_sheet(wb: Workbook, combos: List[dict]):
    """待报列表 sheet，列出本次生成的结算单元组合。"""
    if not combos:
        return
    ws = wb.create_sheet("报税结算单元")
    headers = ["结算单元ID", "结算单元名称", "所属月份", "发放月份", "批次", "人数", "合计收入", "经办人"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, c in enumerate(combos, 2):
        ws.cell(row=row_idx, column=1, value=c.get("unit", ""))
        ws.cell(row=row_idx, column=2, value=c.get("unit_name", ""))
        ws.cell(row=row_idx, column=3, value=c.get("salary_month", ""))
        ws.cell(row=row_idx, column=4, value=c.get("pay_month", ""))
        ws.cell(row=row_idx, column=5, value=c.get("seq", ""))
        ws.cell(row=row_idx, column=6, value=c.get("person_count", ""))
        ws.cell(row=row_idx, column=7, value=c.get("total_income", ""))
        ws.cell(row=row_idx, column=8, value=c.get("handler", ""))
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(25, len(str(h)) * 1.5))


def generate_formula_explanation_sheet(wb: Workbook, records: List[SalaryRecord]):
    """验算公式说明 sheet，逐项解释左=右校验。"""
    ws = wb.create_sheet("验算公式说明")
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55

    def w(row, col, text, bold=False):
        cell = ws.cell(row=row, column=col, value=text)
        if isinstance(text, str) and text.startswith("="):
            cell.data_type = "s"
        if bold:
            cell.font = cell.font.copy(bold=True)
        return cell

    w(1, 1, "验证报告验算公式说明", bold=True)
    w(2, 1, "验证报告 sheet 中每一行对一个人的工资数据进行左=右校验：左与右的差值绝对值小于 0.01 即为通过。")

    w(4, 1, "一、本期收入（报税口径）", bold=True)
    w(5, 1, "本期收入 = 本次工资总额(ATC93AA) − 本次免税(ATC936) − 大病险个人(ATC93BD) − 补缴及退款保险差额个人(ATC93BE)")
    w(6, 1, "说明：本次免税(ATC936) = 采暖费(ATC93W21) + 独生子女费(ATC93W4)，经数据库验证 100% 吻合。")
    w(7, 1, "填报口径：本期收入已扣除本次免税部分，本期免税收入列不填写（业务要求，避免税务核查）。")

    w(8, 1, "二、左(收入-五险一金)", bold=True)
    w(9, 1, "左 = 本期收入 − 当月养老个人缴(BAA001) − 当月失业个人缴(BAA003) − 当月医疗个人缴(BAA002)")
    w(10, 1, "     − 个人公积金月缴存额(CAA002) − 个人其他调整(ATC93AG) − 个人欠款(ATC93E) − 扣款-大病险(ATC93Y2)")
    w(11, 1, "含义：本期收入扣除个人承担的五险一金及其他个人扣减项后，理论上应等于个人实际到手的金额。")

    w(13, 1, "三、右(实发+个税-免税)", bold=True)
    w(14, 1, "右 = 本次实发金额合计(ATC93C) + 税后扣除工会会费(ATC93Z2) + 个人承担代理费(BAA300)")
    w(15, 1, "     + 本次个人所得税(ATC93D) − 免税")
    w(16, 1, "免税 = 本次免税(ATC936) + 大病险个人(ATC93BD)")
    w(17, 1, "通俗理解（为什么这样加减）：")
    w(18, 1, "实发金额是【扣掉工会会费和个税之后】的钱——工会会费(ATC93Z2)和个税(ATC93D)已经从工资里扣走了，")
    w(19, 1, "所以实发金额里【不含】这两项。但左侧的收入没扣这两项，为了两边对齐，右侧要把它们【加回来】。")
    w(20, 1, "实发金额是【含免税】的钱——采暖费、独生子女费、大病险这些免税的钱是发到手的，含在实发金额里。")
    w(21, 1, "但左侧的收入已经把免税部分【减掉】了（免税不计收入），所以右侧也要把免税【减去】，两边才对得上。")
    w(22, 1, "一句话：实发里没扣的两项（工会会费、个税）加回来，实发里含但收入里没算的（免税）减掉。")
    w(23, 1, "举例：某人工资总额 5000（其中含采暖费 500 免税），五险 800，个税 100，工会会费 0。")
    w(24, 1, "  左 = 本期收入(5000−500=4500) − 800 = 3700")
    w(25, 1, "  实发 = 5000 − 800 − 100 = 4100（采暖费 500 已发到手里，含在实发里）")
    w(26, 1, "  右 = 实发(4100) + 工会会费(0) + 个税(100) − 免税(500) = 3700")
    w(27, 1, "  左 = 右 = 3700 ✓ 对上！")
    w(28, 1, "  如果不加个税：4100 − 500 = 3600 ≠ 3700；如果不减免税：4100 + 100 = 4200 ≠ 3700。")
    w(29, 1, "  所以：实发里没扣的两项（工会会费、个税）必须加回，实发里含但收入里不算的（免税）必须减掉。")

    w(27, 1, "四、正确理解（简洁版）", bold=True)
    w(28, 1, "左 = 应计入报税的收入 − 个人五险一金等扣减")
    w(29, 1, "右 = 实际发放的钱 + 代扣的个税 − 已从收入中剔除的免税部分")
    w(30, 1, "两边从不同角度还原同一笔工资，应当相等；不相等则数据有疑点，需人工核对。")

    w(32, 1, "五、字段关系图", bold=True)
    rel_rows = [
        ("", "【收入侧】", ""),
        ("本次工资总额(ATC93AA)", "− 本次免税(ATC936)", "采暖费+独生子女费，免税不计收入"),
        ("", "− 大病险个人(ATC93BD)", "大病险个人部分同样免税"),
        ("", "− 补缴退款差额(ATC93BE)", "补缴/退款冲抵"),
        ("= 本期收入", "← 报税口径的收入", ""),
        ("本期收入", "− 养老(BAA001) − 失业(BAA003) − 医疗(BAA002) − 公积金(CAA002)", "五险一金个人缴"),
        ("", "− 其他调整(ATC93AG) − 个人欠款(ATC93E) − 扣款大病险(ATC93Y2)", "其他扣减"),
        ("= 左", "← 理论到手的钱", ""),
        ("", "", ""),
        ("", "【发放侧】", ""),
        ("本次实发金额(ATC93C)", "+ 税后工会会费(ATC93Z2)", "税后另扣的工会费，还原"),
        ("", "+ 个人代理费(BAA300)", "个人承担的代理费，还原"),
        ("", "+ 个人所得税(ATC93D)", "代扣的个税，还原"),
        ("", "− 免税(ATC936+ATC93BD)", "收入里已剔除的免税部分"),
        ("= 右", "← 从实发反推的同一笔钱", ""),
        ("", "", ""),
        ("左 = 右（差值 < 0.01）", "→ 数据正确", ""),
    ]
    for i, (c1, c2, c3) in enumerate(rel_rows, 33):
        w(i, 1, c1)
        w(i, 2, c2)
        w(i, 3, c3)

    w(51, 1, "六、验证报告列字段对照", bold=True)
    headers = ["验证列", "公式", "涉及字段"]
    for col, h in enumerate(headers, 1):
        w(52, col, h, bold=True)
    rows = [
        ("本期收入", "工资总额 − 本次免税 − 大病险个人 − 补缴及退款保险差额个人", "ATC93AA, ATC936, ATC93BD, ATC93BE"),
        ("左", "本期收入 − 养老 − 失业 − 医疗 − 公积金 − 个人其他调整 − 个人欠款 − 扣款大病险 − 意外险个人", "BAA001, BAA003, BAA002, CAA002, ATC93AG, ATC93E, ATC93Y2, ATC93BH"),
        ("右", "实发合计 + 税后工会会费 + 个人代理费 + 个税 − 免税", "ATC93C, ATC93Z2, BAA300, ATC93D, ATC936, ATC93BD"),
        ("差值", "|左 − 右|", ""),
        ("状态", "差值 < 0.01 为通过", ""),
    ]
    for i, (c1, c2, c3) in enumerate(rows, 53):
        w(i, 1, c1)
        w(i, 2, c2)
        w(i, 3, c3)

    w(59, 1, "七、验算示例（第一条记录实际数值）", bold=True)
    if records:
        rec = records[0]
        income = calc_本期收入(rec)
        tax_exempt = calc_免税(rec)
        left = (income - rec.养老个人 - rec.失业个人 - rec.医疗个人 - rec.公积金个人
                - rec.个人其他调整 - rec.个人欠款 - rec.扣款大病险 - rec.意外险个人)
        right = (rec.实发工资 + rec.税后工会会费 + rec.个人代理费 + rec.个人所得税 - tax_exempt)
        ex_rows = [
            ("姓名", rec.姓名, ""),
            ("工资总额(ATC93AA)", rec.工资总额, ""),
            ("本次免税(ATC936)", rec.补发3, ""),
            ("大病险个人(ATC93BD)", rec.大病险个人, ""),
            ("补缴及退款保险差额个人(ATC93BE)", rec.补缴及退款保险金额个人, ""),
            ("本期收入", income, "工资总额 − 本次免税 − 大病险 − 补缴退款差额"),
            ("本期免税收入", tax_exempt, "本次免税(ATC936) + 大病险个人(ATC93BD)，填报时列不填"),
            ("五险一金(养老+失业+医疗+公积金)", rec.养老个人 + rec.失业个人 + rec.医疗个人 + rec.公积金个人, "BAA001+BAA003+BAA002+CAA002"),
            ("个人其他调整(ATC93AG)", rec.个人其他调整, ""),
            ("个人欠款(ATC93E)", rec.个人欠款, ""),
            ("扣款大病险(ATC93Y2)", rec.扣款大病险, ""),
            ("左", left, "本期收入 − 五险一金 − AG − E − Y2"),
            ("实发合计(ATC93C)", rec.实发工资, ""),
            ("税后工会会费(ATC93Z2)", rec.税后工会会费, ""),
            ("个人代理费(BAA300)", rec.个人代理费, ""),
            ("个人所得税(ATC93D)", rec.个人所得税, ""),
            ("免税", tax_exempt, "本次免税(ATC936) + 大病险个人(ATC93BD)"),
            ("右", right, "实发 + 工会会费 + 代理费 + 个税 − 免税"),
            ("差值", abs(left - right), "|左 − 右|，<0.01 通过"),
        ]
        for i, (name, val, note) in enumerate(ex_rows, 60):
            w(i, 1, name)
            w(i, 2, val)
            w(i, 3, note)

    w(80, 1, "八、东软数据库月份字段说明（重要）", bold=True)
    w(81, 1, "东软系统存在三个易混淆的月份字段，务必区分：")
    w(82, 1, "  ATC931 工资所属年月   — 工资属于哪个月（如6月工资=202606）")
    w(83, 1, "  ATC932 工资发放年月   — 工资实际发放的月份（TC93表内字段）")
    w(84, 1, "  ATC8G7 经办年月       — 发放年月的【最终依据】（TC8M表字段，本工具'发放月份'下拉即此字段）")
    w(85, 1, "本工具选择'发放月份'时，实际筛选的是 TC8M.ATC8G7（经办年月）。")
    w(86, 1, "同一组合(结算单元+所属月份+批次)下，TC93.ATC932 可能与 TC8M.ATC8G7 不一致：")
    w(87, 1, "如 单元37185 所属202606批1 的 TC93.ATC932=202606（6月发放），但 TC8M.ATC8G7=202607（7月经办申报）。")
    w(88, 1, "因此 TC93总表 sheet 中会出现'发放年月(ATC932)=202606'而您选择发放月份202607的记录——这是正常的，")
    w(89, 1, "请以 TC93总表 最后列的'经办年月(ATC8G7)'为准判断该笔工资属于哪个申报期。")


def generate_raw_detail_sheet(wb: Workbook, raw_records: List[SalaryRecord],
                              combo_map: dict, title: str):
    """原始明细(未合并)sheet：逐条列出未合并记录的报送数据与验算，保证可追溯。"""
    ws = wb.create_sheet("原始明细(未合并)")
    headers = [
        "ATC930", "姓名", "证件号码", "结算单元", "所属月份", "批次",
        "工资总额(AA)", "本次免税(936)", "大病险(BD)", "补缴退款差额(BE)", "本期收入",
        "养老(BAA001)", "失业(BAA003)", "医疗(BAA002)", "公积金(CAA002)",
        "其他调整(AG)", "个人欠款(E)", "扣款大病险(Y2)", "意外险(BH)",
        "实发(C)", "工会会费(Z2)", "代理费(BAA300)", "个税(D)", "免税",
        "左", "右", "差值", "状态", "备注"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for idx, rec in enumerate(raw_records, 2):
        income = calc_本期收入(rec)
        tax_exempt = calc_免税(rec)
        left = (income - rec.养老个人 - rec.失业个人 - rec.医疗个人 - rec.公积金个人
                - rec.个人其他调整 - rec.个人欠款 - rec.扣款大病险 - rec.意外险个人)
        right = (rec.实发工资 + rec.税后工会会费 + rec.个人代理费 + rec.个人所得税 - tax_exempt)
        diff = abs(left - right)
        remark = combo_map.get((rec.结算单元, rec.工资所属年月, rec.当月批次), title)
        vals = [
            rec.tc930_id, rec.姓名, rec.身份证, rec.结算单元, rec.工资所属年月, rec.当月批次,
            rec.工资总额, rec.补发3, rec.大病险个人, rec.补缴及退款保险金额个人, income,
            rec.养老个人, rec.失业个人, rec.医疗个人, rec.公积金个人,
            rec.个人其他调整, rec.个人欠款, rec.扣款大病险, rec.意外险个人,
            rec.实发工资, rec.税后工会会费, rec.个人代理费, rec.个人所得税, tax_exempt,
            left, right, round(float(diff), 4), "通过" if diff < 0.01 else "失败", remark
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=idx, column=col, value=val)


def generate_merge_detail_sheet(wb: Workbook, raw_records: List[SalaryRecord],
                                merged_records: List[SalaryRecord]):
    """合并明细sheet：按人+所属月份分组，展示原始条目与合并后汇总，合并过程可追溯。"""
    ws = wb.create_sheet("合并明细")
    groups = {}
    for rec in raw_records:
        groups.setdefault((rec.职工号, rec.工资所属年月), []).append(rec)

    headers = [
        "姓名", "证件号码", "所属月份", "原始条数",
        "原始记录(ATC930/本期收入/个税)", "合并后结算单元", "合并后批次",
        "合并本期收入", "合并五险", "合并个税", "合并实发", "合并免税",
        "合并左", "合并右", "差值", "状态"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    merged_map = {(m.职工号, m.工资所属年月): m for m in merged_records}
    row = 2
    for (pid, month), recs in sorted(groups.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        m = merged_map.get((pid, month))
        raw_desc = "; ".join(
            f"{r.tc930_id}/{round(float(calc_本期收入(r)), 2)}/{r.个人所得税}" for r in recs)
        if m is None:
            continue
        income = calc_本期收入(m)
        tax_exempt = calc_免税(m)
        left = (income - m.养老个人 - m.失业个人 - m.医疗个人 - m.公积金个人
                - m.个人其他调整 - m.个人欠款 - m.扣款大病险 - m.意外险个人)
        right = (m.实发工资 + m.税后工会会费 + m.个人代理费 + m.个人所得税 - tax_exempt)
        diff = abs(left - right)
        vals = [
            recs[0].姓名, recs[0].身份证, month, len(recs),
            raw_desc, m.结算单元, m.当月批次,
            income, m.养老个人 + m.失业个人 + m.医疗个人 + m.公积金个人, m.个人所得税,
            m.实发工资, tax_exempt,
            left, right, round(float(diff), 4), "通过" if diff < 0.01 else "失败"
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
        row += 1