"""人员信息采集导入模板生成器 - 精确复制 demo TemplateFiller.cs 算法"""
from datetime import datetime
from typing import List
import os
from openpyxl import Workbook
from models import PersonnelInfo, GenerateResult
from templates_gen.explanation import add_explanation_sheet


def extract_remark(title: str) -> str:
    """从标题提取备注字段"""
    if not title:
        return ""
    t = title.replace("东北师范大学人事处", "").replace("劳务派遣人员工资发放表", "").strip()
    t = t.replace("年", "").replace("月", "").replace("系统", "").strip()
    digits = ''.join(c for c in t if c.isdigit())
    if len(digits) >= 4:
        t = t.replace(digits, "").strip()
    return t if t else title


def get_gender_from_id(id_card: str) -> str:
    """从身份证第17位提取性别（奇=男，偶=女）"""
    if not id_card or len(id_card) < 18:
        return ""
    try:
        d = int(id_card[16])
        return "男" if d % 2 == 1 else "女"
    except (ValueError, IndexError):
        return ""


def get_birth_date_from_id(id_card: str) -> str:
    """从身份证第7-14位提取出生日期（YYYYMMDD）"""
    if not id_card or len(id_card) < 14:
        return ""
    return id_card[6:14]


def generate_personnel_info(personnel: List[PersonnelInfo], title: str, output_dir: str) -> GenerateResult:
    """生成人员信息采集导入模板 Excel"""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(output_dir, f"人员信息采集导入模板_{timestamp}.xlsx")
    remark = extract_remark(title)
    
    # 按职工号去重
    seen = set()
    unique_personnel = []
    for p in personnel:
        if p.职工号 not in seen:
            seen.add(p.职工号)
            unique_personnel.append(p)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "人员信息"
    
    # 51列标题 - 必须与 demo 完全一致
    headers = [
        "工号", "*姓名", "*证件类型", "*证件号码", "*国籍(地区)", "*性别",
        "*出生日期", "是否高级专家", "*任职受雇从业类型", "其他情况说明",
        "入职年度就业情形", "手机号码", "任职受雇从业日期", "离职日期",
        "是否离职后补发工资", "实际补发工资的月份", "是否残疾", "是否烈属",
        "是否孤老", "残疾证件类型", "残疾证号", "烈属证号", "是否扣除减除费用",
        "个人投资额", "个人投资比例(%)", "备注", "中文名", "涉税事由",
        "出生国家(地区)", "首次入境时间", "预计离境时间", "其他证件类型",
        "其他证件号码", "户籍所在地（省）", "户籍所在地（市）", "户籍所在地（区县）",
        "户籍所在地（详细地址）", "经常居住地（省）", "经常居住地（市）",
        "经常居住地（区县）", "经常居住地（详细地址）", "联系地址（省）",
        "联系地址（市）", "联系地址（区县）", "联系地址（详细地址）",
        "电子邮箱", "学历", "开户银行", "银行账号", "开户行省份", "职务"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # 写入数据
    for idx, p in enumerate(unique_personnel, 1):
        row = idx + 1
        gender = get_gender_from_id(p.身份证)
        birth_date = get_birth_date_from_id(p.身份证)
        
        ws.cell(row=row, column=1, value=p.职工号)
        ws.cell(row=row, column=2, value=p.姓名)
        ws.cell(row=row, column=3, value="居民身份证")
        ws.cell(row=row, column=4, value=p.身份证)
        ws.cell(row=row, column=5, value="中国")
        ws.cell(row=row, column=6, value=gender)
        if birth_date:
            ws.cell(row=row, column=7, value=birth_date)
        ws.cell(row=row, column=9, value="雇员")
        ws.cell(row=row, column=26, value=remark)
    
    add_explanation_sheet(wb, [
        ("人员信息", [
            "51 列个税人员信息采集导入模板，无校验。",
            "按职工号去重，每人一行。",
            "身份证解析：*性别 = 身份证第 17 位奇男偶女；*出生日期 = 第 7~14 位(YYYYMMDD)。",
            "*国籍(地区) 恒为「中国」；*证件类型 恒为「居民身份证」；*任职受雇从业类型 恒为「雇员」。",
            "备注从标题中提取（去除机构名/年月/数字后剩余文本）。",
        ]),
    ])
    
    wb.save(output_path)
    
    return GenerateResult(
        file_path=output_path,
        template_type="人员信息采集导入模板",
        record_count=len(unique_personnel),
        validation_pass=0,
        validation_fail=0
    )
