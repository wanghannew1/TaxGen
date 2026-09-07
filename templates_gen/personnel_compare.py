"""增减员比对 - 51列模板列头定义与数据映射

将个税端导出人员信息与当月发薪人员比对后, 增员/减员名单均以
人员信息采集导入模板的 51 列格式输出 (与 personnel_info.py 的
列头完全一致)。

列号索引 (0-based, 对应 51 列):
- 0: 工号, 1: *姓名, 2: *证件类型, 3: *证件号码, 4: *国籍(地区), 5: *性别
- 6: *出生日期, 7: 是否高级专家, 8: *任职受雇从业类型, 9: 其他情况说明
- 10: 入职年度就业情形, 11: 手机号码, 12: 任职受雇从业日期, 13: 离职日期
- 14: 是否离职后补发工资, ..., 25: 备注, ..., 50: 职务
"""
from datetime import datetime
from typing import List

from openpyxl import Workbook

from models import GenerateResult, PersonnelInfo
from templates_gen.explanation import add_explanation_sheet

# 51 列标题 - 与 personnel_info.py 完全一致
COMPARE_HEADERS: List[str] = [
    "工号", "*姓名", "*证件类型", "*证件号码", "*国籍(地区)", "*性别",
    "*出生日期", "是否高级专家", "*任职受雇从业类型", "其他情况说明",
    "入职年度就业情形", "手机号码", "任职受雇从业日期", "离职日期",
    "是否离职后补发工资", "实际补发工资的月份", "是否残疾", "是否烈属",
    "是否孤老", "残疾证件类型", "残疾证号", "烈属证号", "是否扣除减除费用",
    "个人投资额", "个人投资比例(%)", "备注", "中文名", "涉税事由",
    "出生国家(地区)", "首次入境时间", "预计离境时间", "其他证件类型",
    "其他证件号码", "户籍所在地（省）", "户籍所在地（市）", "户籍所在地（区县）",
    "户籍所在地（详细地址）", "经常居住地（省）", "经常居住地（市）",
    "经常居住地（区县）", "经常居住地（详细地址）", "联系地址（省）",
    "联系地址（市）", "联系地址（区县）", "联系地址（详细地址）",
    "电子邮箱", "学历", "开户银行", "银行账号", "开户行省份", "职务"
]

# 常用字段的列索引 (0-based)
IDX_工号 = 0
IDX_姓名 = 1
IDX_证件类型 = 2
IDX_证件号码 = 3
IDX_国籍 = 4
IDX_性别 = 5
IDX_出生日期 = 6
IDX_任职受雇从业类型 = 8
IDX_手机号码 = 11
IDX_任职受雇从业日期 = 12
IDX_离职日期 = 13


def _normalize_date(value) -> str:
    """将日期值规范化为 "YYYY-MM-DD" 字符串。

    支持 datetime 对象、xlrd 日期序列号 (float)、或已有字符串。
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and not value.is_integer():
        try:
            import xlrd
            return xlrd.xldate_as_datetime(value, 0).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            return str(value)
    s = str(value).strip()
    # 兼容 "1981-10-26 00:00:00" 截断为日期
    if len(s) > 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s


def map_personnel_info_to_row(person: PersonnelInfo) -> list:
    """将数据库人员信息 (增员) 映射为 51 列行数据。"""
    row = [""] * len(COMPARE_HEADERS)
    row[IDX_工号] = person.职工号
    row[IDX_姓名] = person.姓名
    row[IDX_证件类型] = person.证件类型 or "居民身份证"
    row[IDX_证件号码] = person.身份证
    row[IDX_国籍] = person.国籍 or "中国"
    row[IDX_性别] = person.性别
    row[IDX_出生日期] = _normalize_date(person.出生日期)
    row[IDX_任职受雇从业类型] = person.任职类型 or "雇员"
    row[IDX_手机号码] = person.手机号码
    row[IDX_任职受雇从业日期] = _normalize_date(person.任职受雇从业日期)
    row[25] = person.备注  # 备注 = 结算单元名称 (ATB931)
    return row


def map_tax_export_to_row(person: dict, termination_date=None) -> list:
    """将个税端导出记录 (减员) 映射为 51 列行数据。

    termination_date 为工资结束日期 (TC90.ATC90AV 工资结束年月补全为月末,
    datetime 或 str), 填入"离职日期"列标记减员。
    """
    row = [""] * len(COMPARE_HEADERS)
    row[IDX_工号] = str(person.get("工号") or "")
    row[IDX_姓名] = str(person.get("姓名") or "")
    row[IDX_证件类型] = str(person.get("证件类型") or "") or "居民身份证"
    row[IDX_证件号码] = str(person.get("证件号码") or "")
    row[IDX_国籍] = str(person.get("国籍") or "") or "中国"
    row[IDX_性别] = str(person.get("性别") or "")
    row[IDX_出生日期] = _normalize_date(person.get("出生日期"))
    row[IDX_任职受雇从业类型] = str(person.get("任职受雇从业类型") or "") or "雇员"
    row[IDX_手机号码] = str(person.get("手机号码") or "")
    row[IDX_任职受雇从业日期] = _normalize_date(person.get("任职受雇从业日期"))
    if termination_date:
        row[IDX_离职日期] = _normalize_date(termination_date)
    return row


def _cert_key(value) -> str:
    return str(value or "").strip().upper()


# 零申报 Sheet: 正常工资薪金所得 29 列 (与 normal_salary.py 一致), 仅 本期收入+五险一金 填 0
ZERO_HEADERS: List[str] = [
    "工号", "*姓名", "*证件类型", "*证件号码", "本期收入", "本期免税收入",
    "基本养老保险费", "基本医疗保险费", "失业保险费", "住房公积金",
    "累计子女教育", "累计继续教育", "累计住房贷款利息", "累计住房租金",
    "累计赡养老人", "累计3岁以下婴幼儿照护", "累计个人养老金", "企业(职业)年金",
    "商业健康保险", "税延养老保险", "公务交通费用", "通讯费用", "律师办案费用",
    "西藏附加减除费用", "其他", "准予扣除的捐赠额", "减免税额", "协定减免", "备注",
]
# 身份证列索引 (0-based)
ZERO_IDX_工号 = 0
ZERO_IDX_姓名 = 1
ZERO_IDX_证件类型 = 2
ZERO_IDX_证件号码 = 3
# 需填 0 的收入/扣款列 (0-based):
# 本期收入(4) + 基本养老保险费(6) + 基本医疗保险费(7) + 失业保险费(8) + 住房公积金(9)
# 其余数值列 (本期免税收入、累计*、企业(职业)年金、减免等) 一律留空, 不填 0。
# 注意: 本期免税收入(5) 不在填 0 之列。
ZERO_IDX_本期收入 = 4
ZERO_IDX_基本养老保险费 = 6
ZERO_IDX_基本医疗保险费 = 7
ZERO_IDX_失业保险费 = 8
ZERO_IDX_住房公积金 = 9
ZERO_ZERO_COLS = [ZERO_IDX_本期收入, ZERO_IDX_基本养老保险费, ZERO_IDX_基本医疗保险费,
                  ZERO_IDX_失业保险费, ZERO_IDX_住房公积金]
# 备注列 (0-based), 单独取值
ZERO_IDX_备注 = 28


def build_zero_declare_row(person: dict) -> list:
    """将个税端导出记录映射为 29 列零申报行。

    在职且本期无工资、也未减员的人员需 0 申报: 仅 本期收入、基本养老保险费、
    基本医疗保险费、失业保险费、住房公积金 填 0, 其余数值列 (本期免税收入、
    累计*、企业(职业)年金、减免税额等) 留空; 仅工号/姓名/证件类型/证件号码
    有值, 备注为结算单元名称。
    """
    row = [""] * len(ZERO_HEADERS)
    row[ZERO_IDX_工号] = str(person.get("工号") or "")
    row[ZERO_IDX_姓名] = str(person.get("姓名") or "")
    row[ZERO_IDX_证件类型] = str(person.get("证件类型") or "") or "居民身份证"
    row[ZERO_IDX_证件号码] = str(person.get("证件号码") or "")
    for i in ZERO_ZERO_COLS:
        row[i] = 0
    row[ZERO_IDX_备注] = str(person.get("备注") or "")  # 备注 = 结算单元名称
    return row


# 增员验证附加列 (51 列之后追加)
VERIFY_HEADERS = [
    # 公共
    "增员原因",
    # 发薪块
    "发薪月份范围",
    "发薪月份有薪资",
    "发薪薪资明细(结算单元-所属月-批次)",
    "发薪本期收入",
    "发薪养老",
    "发薪医疗",
    "发薪失业",
    "发薪公积金",
    "发薪经办人",
    # 未发薪块
    "未发薪所属月份范围",
    "有未发薪工资表",
    "未发薪明细(结算单元-所属月-批次)",
    "未发薪本期收入",
    "未发薪基本养老保险费",
    "未发薪基本医疗保险费",
    "未发薪失业保险费",
    "未发薪住房公积金",
    "未发薪经办人",
    # 合同块
    "合同签署时间范围",
    "合同签署增员",
    "合同开始日期",
    "合同经办人",
    # 公共
    "是否零申报",
]

# 减员验证附加列 (51 列之后追加)
REMOVE_VERIFY_HEADERS = [
    "减员类型",
    # 个税端信息
    "个税端姓名",
    "个税端证件类型",
    "个税端证件号码",
    "个税端报送状态",
    "身份验证状态",
    "任职受雇从业日期(个税端)",
    "任职受雇从业类型",
    "个税端离职日期",
    "其他情况说明",
    "备注",
    "更新时间",
    "手机号码",
    # 发薪块
    "发薪月份范围",
    "发薪月份有薪资",
    "发薪薪资明细(结算单元-所属月-批次)",
    "最后一次发薪(结算单元-所属月-批次)",
    # 未发薪块
    "未发薪所属月份范围",
    "有未发薪工资表",
    "未发薪明细(结算单元-所属月-批次)",
    # 合同块 (Oracle)
    "合同签署时间范围",
    "合同签署增员",
    "合同开始日期（ATC90C）",
    "合同终止日期（ATC90D）",
    "工资结束年月（ATC90AV）",
    "合同经办人",
]

# 明细 Sheet 列头
TC93_HEADERS = ["证件号码", "姓名", "结算单元代码", "结算单元名称", "所属月份", "批次",
                "应发工资", "本期收入", "养老", "医疗", "失业", "公积金"]
TC8M_HEADERS = ["证件号码", "姓名", "结算单元代码", "结算单元名称", "发放月份", "所属月份", "批次", "经办人"]
TC90_HEADERS = ["证件号码", "姓名", "结算单元代码", "结算单元名称", "合同开始日期", "合同终止日期", "单位名称", "经办人"]


def build_verify_row(add_row, cert, params, paid_salary_details, unpaid_salary_details, tc8m_details, contract_start, contract_details=None):
    """为单个增员人员组装增员验证行。

    Args:
        add_row: 51 列增员行 (作为前 51 列)
        cert: 证件号码
        params: {pay_months, unpaid_months, contract_start, contract_end}
        paid_salary_details: 该人发薪对应的 TC93 工资记录 (按发薪记录所属月份)
        unpaid_salary_details: 该人未发薪范围的 TC93 工资记录 (所属月份范围)
        tc8m_details: 该人 TC8M 发放记录列表 (发薪月份范围)
        contract_start: 该人合同开始日期 (或 None)
        contract_details: 该人 TC90 合同记录列表 (取经办人)

    Returns:
        51 + len(VERIFY_HEADERS) 列的行数据
    """
    contract_details = contract_details or []
    # 发薪: 发薪月份范围有 TC8M 已发记录 (按 所属月-批次 去重)
    paid = []
    paid_key_set = set()
    for t in tc8m_details:
        if t.get("pay_month") not in params["pay_months"]:
            continue
        key = (t.get("salary_month"), t.get("seq"))
        if key in paid_key_set:
            continue
        paid_key_set.add(key)
        paid.append(t)
    has_paid = bool(paid)
    # 发薪对应的工资明细: 只取该人 TC8M 已发 (salary_month, seq) 对应的 TC93 记录
    paid_keys = {(t.get("salary_month"), t.get("seq")) for t in paid}
    paid_salary = []
    seen_paid = set()
    for r in paid_salary_details:
        key = (r["salary_month"], r["seq"])
        if key not in paid_keys or key in seen_paid:
            continue
        seen_paid.add(key)
        paid_salary.append(r)
    paid_income = sum(r["本期收入"] for r in paid_salary)
    # 未发薪: 所属月份范围有 TC93 记录, 但不在任何发薪月份的 TC8M 已发中
    paid_set = {(t.get("salary_month"), t.get("seq")) for t in tc8m_details}
    unpaid = [r for r in unpaid_salary_details
              if r.get("salary_month") in params["unpaid_months"]
              and (r["salary_month"], r["seq"]) not in paid_set]
    has_unpaid = bool(unpaid)
    # 合同签署: 无发薪且无未发薪, 但合同开始日期在合同签署时间范围
    has_contract = bool(contract_start) and not has_paid and not has_unpaid
    # 零申报: 发薪月份无薪资 (未发薪/合同签署均属零申报)
    zero = not has_paid

    reasons = []
    if has_paid:
        reasons.append("发薪")
    if has_unpaid:
        reasons.append("未发薪")
    if has_contract:
        reasons.append("合同")
    reason_str = "+".join(reasons)

    # 经办人: 发薪取 TC8M 已发记录的 AAE019, 未发薪取 TC93 记录 AAE019,
    # 合同取 TC90 记录 AAE019 (均去重后分号连接)
    paid_handlers = "; ".join(sorted({str(t.get("handler") or "") for t in paid if t.get("handler")}))
    unpaid_handlers = "; ".join(sorted({str(r.get("handler") or "") for r in unpaid if r.get("handler")}))
    contract_handlers = "; ".join(sorted(
        {str(c.get("handler") or c.get("经办人") or "") for c in contract_details
         if c.get("handler") or c.get("经办人")}))

    contract_time = ""
    if params.get("contract_start") or params.get("contract_end"):
        contract_time = f"{params.get('contract_start')} ~ {params.get('contract_end')}"

    return add_row + [
        reason_str,
        # 发薪块
        "/".join(str(m) for m in params["pay_months"]),
        "是" if has_paid else "否",
        "; ".join(f"{t['unit_name'] or t['unit_code']}-{t['salary_month']}-{t['seq']}" for t in paid),
        round(paid_income, 2),
        round(sum(r["养老"] for r in paid_salary), 2),
        round(sum(r["医疗"] for r in paid_salary), 2),
        round(sum(r["失业"] for r in paid_salary), 2),
        round(sum(r["公积金"] for r in paid_salary), 2),
        paid_handlers,
        # 未发薪块
        "/".join(str(m) for m in params["unpaid_months"]),
        "是" if has_unpaid else "否",
        "; ".join(f"{r['unit_name'] or r['unit_code']}-{r['salary_month']}-{r['seq']}" for r in unpaid),
        round(sum(r["本期收入"] for r in unpaid), 2),
        round(sum(r["养老"] for r in unpaid), 2),
        round(sum(r["医疗"] for r in unpaid), 2),
        round(sum(r["失业"] for r in unpaid), 2),
        round(sum(r["公积金"] for r in unpaid), 2),
        unpaid_handlers,
        # 合同块
        contract_time,
        "是" if has_contract else "否",
        contract_start or "",
        contract_handlers,
        # 公共
        "是" if zero else "否",
    ]


def build_remove_verify_row(remove_row, cert, remove_type, params,
                            paid_salary_details, unpaid_salary_details, tc8m_details,
                            contract_start, contract_end_dt, contract_end_d90=None,
                            contract_details=None,
                            tax_person=None, last_pay=None):
    """为单个减员人员组装减员验证行。

    减员人员应不在发薪/未发薪/当期合同签署名单中, 各来源块验证列应为否/空,
    用于确认减员判断正确。TC90 合同信息为减员判断依据。

    Args:
        remove_row: 51 列减员行 (来自个税端导出, 作为前 51 列)
        cert: 证件号码
        remove_type: 减员类型 (近期离职/待确认近期离职)
        params: {pay_months, unpaid_months, contract_start, contract_end}
        paid_salary_details: 该人 TC93 发薪工资记录 (验证用, 应为空)
        unpaid_salary_details: 该人 TC93 未发薪记录 (验证用, 应为空)
        tc8m_details: 该人 TC8M 发放记录 (验证用, 应为空)
        contract_start: 该人合同开始日期 (TC90 最早 ATC90C)
        contract_end_dt: 该人工资结束年月 (TC90 ATC90AV 补全月末, 判定离职口径)
        contract_end_d90: 该人合同终止日期 (TC90 ATC90D)
        contract_details: 该人 TC90 合同记录列表 (取经办人)
        tax_person: 个税端导出原始记录 dict (报送状态/身份验证/任职/更新时间等)
        last_pay: 该人最近一次发薪记录 dict {unit_name, salary_month, seq, pay_month}

    Returns:
        51 + len(REMOVE_VERIFY_HEADERS) 列的行数据
    """
    contract_details = contract_details or []
    tax_person = tax_person or {}
    # 减员人员应无发薪/未发薪记录 (验证其不在保护名单)
    has_paid = bool(tc8m_details)
    has_unpaid = bool(unpaid_salary_details)
    paid_detail = "; ".join(
        f"{t['unit_name'] or t['unit_code']}-{t['salary_month']}-{t['seq']}" for t in tc8m_details)
    unpaid_detail = "; ".join(
        f"{r['unit_name'] or r['unit_code']}-{r['salary_month']}-{r['seq']}" for r in unpaid_salary_details)
    contract_handlers = "; ".join(sorted(
        {str(c.get("handler") or c.get("经办人") or "") for c in contract_details
         if c.get("handler") or c.get("经办人")}))
    last_pay_detail = ""
    if last_pay:
        last_pay_detail = f"{last_pay.get('unit_name', '')}-{last_pay.get('salary_month', '')}-{last_pay.get('seq', '')}"

    contract_time = ""
    if params.get("contract_start") or params.get("contract_end"):
        contract_time = f"{params.get('contract_start')} ~ {params.get('contract_end')}"

    return remove_row + [
        remove_type,
        # 个税端信息
        str(tax_person.get("姓名") or ""),
        str(tax_person.get("证件类型") or ""),
        str(tax_person.get("证件号码") or ""),
        str(tax_person.get("报送状态") or ""),
        str(tax_person.get("身份验证状态") or ""),
        str(tax_person.get("任职受雇从业日期") or ""),
        str(tax_person.get("任职受雇从业类型") or ""),
        str(tax_person.get("离职日期") or ""),
        str(tax_person.get("其他情况说明") or ""),
        str(tax_person.get("备注") or ""),
        str(tax_person.get("更新时间") or ""),
        str(tax_person.get("手机号码") or ""),
        # 发薪块
        "/".join(str(m) for m in params["pay_months"]),
        "是" if has_paid else "否",
        paid_detail,
        last_pay_detail,
        # 未发薪块
        "/".join(str(m) for m in params["unpaid_months"]),
        "是" if has_unpaid else "否",
        unpaid_detail,
        # 合同块 (Oracle)
        contract_time,
        "否",
        contract_start or "",
        contract_end_d90.strftime("%Y-%m-%d") if hasattr(contract_end_d90, "strftime") else (contract_end_d90 or ""),
        contract_end_dt.strftime("%Y-%m-%d") if hasattr(contract_end_dt, "strftime") else (contract_end_dt or ""),
        contract_handlers,
    ]


def compare_personnel(tax_export_persons, payroll_certs, payroll_personnel, salary_end_dates,
                      unpaid_persons=None, contract_signed_persons=None,
                      person_units=None, filter_handlers=None, filter_units=None,
                      filter_depts=None, exclude_certs=None,
                      unpaid_latest_persons=None,
                      payroll_start_certs=None,
                      deferred_pay_persons=None):
    """增减员比对引擎。

    Args:
        tax_export_persons: List[dict] - 个税端导出文件解析结果
        payroll_certs: Set[str] - 发薪月份范围内发薪人员证件号集合 (已大写, 并集)
        payroll_personnel: List[PersonnelInfo] - 发薪人员详细信息
        salary_end_dates: Dict[str, datetime] - 证件号 → 工资结束日期
            (TC90.ATC90AV 工资结束年月补全为月末, 即离职日期)
        unpaid_persons: Set[str] - 未发薪工资表人员证件号集合 (减员排除 + 增员条件②)
        contract_signed_persons: Set[str] - 合同签署时间范围内人员 (增员条件③)
        person_units: Dict[str, dict] - 证件号 → {pay_handlers, salary_handlers,
            contract_handlers, last_pay_handlers, last_salary_handlers,
            unit_code, unit_name, dept_name}
        filter_handlers: List[str] - 经办人过滤 (空=不过滤)
        filter_units: List[int] - 结算单元代码过滤 (空=不过滤)
        filter_depts: List[str] - 单位名称过滤 (空=不过滤)
        exclude_certs: Set[str] - 需从增员中排除的人员 (特殊结算单元工资为0)
        unpaid_latest_persons: Set[str] - 最后一笔工资未发放的人员
            (压月发薪延期: 最后一笔工资未发完前不能减员, 从减员候选剔除)
        payroll_start_certs: Set[str] - 申报期(发薪起始月)发薪人员证件号集合。
            零申报人群判定用: 个税端在职且申报期无工资、也未减员 → 需 0 申报。
        deferred_pay_persons: Set[str] - 延期发放人员证件号集合 (次月月初 1~10 日
            有 TC8M 已发记录, 证明所选申报月在职)。从"待确认近期离职"中剔除,
            改判零申报 (不参与待确认零申报)。

    Returns:
        (add_rows, departed_rows, pending_rows, stats):
            add_rows: List[list] - 增员 51 列行数据
            departed_rows: List[list] - 近期离职人员 51 列行数据
            pending_rows: List[list] - 待确认近期离职人员 51 列行数据
            stats: dict - {add_count, departed_count, pending_count, tax_total, payroll_total,
                zero_count, zero_certs, zero_pending_count, zero_pending_certs,
                active_total, unpaid_total, contract_total, filtered_active_count, filtered_payroll_count}
    """
    unpaid_set = {_cert_key(c) for c in (unpaid_persons or set())}
    unpaid_set.discard("")
    contract_set = {_cert_key(c) for c in (contract_signed_persons or set())}
    contract_set.discard("")
    exclude_set = {_cert_key(c) for c in (exclude_certs or set())}
    exclude_set.discard("")
    unpaid_latest_set = {_cert_key(c) for c in (unpaid_latest_persons or set())}
    unpaid_latest_set.discard("")
    payroll_start_set = {_cert_key(c) for c in (payroll_start_certs or set())}
    payroll_start_set.discard("")

    handlers = {h for h in (filter_handlers or []) if h}
    units = {int(u) for u in (filter_units or []) if u}
    depts = {d for d in (filter_depts or []) if d}
    person_units = person_units or {}

    def _passes_filter(cert):
        if not handlers and not units and not depts:
            return True
        info = person_units.get(_cert_key(cert))
        if not info:
            return False
        if handlers:
            # 有效经办人 E (按身份类别取第一个存在的来源):
            # 增员: 发薪(所选月TC8M) → 做工资(未发薪月TC93) → 合同(最后一份TC90)
            # 减员: 最后一次发薪(TC8M) → 最后一次做工资(TC93) → 最后一份合同(TC90)
            if cert in payroll_set:
                e = info.get("pay_handlers") or []
            elif cert in unpaid_set:
                e = info.get("salary_handlers") or []
            elif cert in contract_set:
                e = info.get("contract_handlers") or []
            else:
                e = (info.get("last_pay_handlers") or info.get("last_salary_handlers")
                     or info.get("contract_handlers") or [])
            if not (set(e) & handlers):
                return False
        if units and info.get("unit_code") not in units:
            return False
        if depts and info.get("dept_name") not in depts:
            return False
        return True

    tax_certs = {_cert_key(p.get("证件号码")) for p in tax_export_persons}
    tax_certs.discard("")

    payroll_set = {_cert_key(c) for c in payroll_certs}
    payroll_set.discard("")

    # 增员 = (发薪 ∪ 未发薪工资表 ∪ 合同签署) - 个税端, 再按经办人/结算单元过滤
    add_certs = (payroll_set | unpaid_set | contract_set) - tax_certs
    add_certs = {c for c in add_certs if _passes_filter(c)}
    add_certs -= exclude_set

    # 疑似离职: 个税端未标记离职(无离职日期)且不在发薪/未发薪名单
    # (个税端已有离职日期的属历史离职, 不参与近期离职判定)
    active_certs = {_cert_key(p.get("证件号码")) for p in tax_export_persons
                    if not str(p.get("离职日期") or "").strip()}
    protected_certs = payroll_set | unpaid_set | contract_set
    # 最后一笔工资未发放的人员不参与减员 (压月发薪延期规则)
    suspect_certs = active_certs - protected_certs - unpaid_latest_set
    suspect_certs = {c for c in suspect_certs if _passes_filter(c)}
    departed_certs = {c for c in suspect_certs if c in salary_end_dates}
    deferred_set = {_cert_key(c) for c in (deferred_pay_persons or set())}
    deferred_set.discard("")
    # 延期发放人员: 申报月无发放但次月月初已发 (在职证据), 未判离职则不列待确认
    pending_certs = suspect_certs - departed_certs - deferred_set

    # 零申报人群 (主 sheet): 个税端在职(无离职日期) 且 申报期(发薪起始月)无工资
    # 且 未减员(departed+pending 均排除), 再排除特殊结算单元(工资为0/完全排除 均不报税)。
    # 延期发放人员因已从 pending 剔除, 会自然落入零申报人群。
    zero_declare = active_certs - payroll_start_set - departed_certs - pending_certs
    zero_declare = zero_declare - exclude_set
    zero_declare = {c for c in zero_declare if _passes_filter(c)}
    # 待确认零申报 (独立 sheet): 待确认人员虽被判为待查, 一旦确认非离职可直接 0 申报,
    # 故单独输出。排除特殊结算单元 (pending 已按经办人过滤, 故仅再滤特殊单元)。
    zero_pending_declare = pending_certs - exclude_set
    zero_pending_declare = {c for c in zero_pending_declare if _passes_filter(c)}

    # 新增 5 个统计字段（纯函数计算，零新增 DB 查询）
    active_total = len(active_certs)  # 个税端在职人数
    unpaid_total = len(unpaid_set)    # 未发薪工资表人数
    contract_total = len(contract_set)  # 合同签署人数
    filtered_active_count = len({c for c in active_certs if _passes_filter(c)})  # 过滤后在职人数
    filtered_payroll_count = len({c for c in payroll_set if _passes_filter(c)})  # 过滤后发薪人数

    add_rows = []
    for person in payroll_personnel:
        if _cert_key(person.身份证) in add_certs:
            add_rows.append(map_personnel_info_to_row(person))

    departed_rows = []
    for person in tax_export_persons:
        cert = _cert_key(person.get("证件号码"))
        if cert in departed_certs:
            departed_rows.append(map_tax_export_to_row(
                person, termination_date=salary_end_dates.get(cert)))

    pending_rows = []
    for person in tax_export_persons:
        cert = _cert_key(person.get("证件号码"))
        if cert in pending_certs:
            pending_rows.append(map_tax_export_to_row(person))

    stats = {
        "add_count": len(add_certs),
        "departed_count": len(departed_certs),
        "pending_count": len(pending_certs),
        "tax_total": len(tax_certs),
        "payroll_total": len(payroll_set),
        "zero_count": len(zero_declare),
        "zero_certs": zero_declare,
        "zero_pending_count": len(zero_pending_declare),
        "zero_pending_certs": zero_pending_declare,
        "add_certs": add_certs,
        # 新增 5 个统计字段（纯函数计算，零新增 DB 查询）
        "active_total": active_total,
        "unpaid_total": unpaid_total,
        "contract_total": contract_total,
        "filtered_active_count": filtered_active_count,
        "filtered_payroll_count": filtered_payroll_count,
    }
    return add_rows, departed_rows, pending_rows, stats


def generate_compare_excel(add_rows, departed_rows, pending_rows, stats, output_dir: str, pay_month: int,
                           verify_rows=None, tc93_rows=None, tc8m_rows=None, tc90_rows=None,
                           remove_verify_rows=None, zero_rows=None, zero_pending_rows=None) -> GenerateResult:
    """生成增减员比对结果 Excel。

    基础三 Sheet: 增员名单 + 近期离职人员 + 待确认近期离职人员。
    可选附加 Sheet: 增员验证 (51列+验证列), 减员验证, TC93工资明细, TC8M发放明细, TC90合同明细,
    零申报 (29列正常工资薪金所得, 收入全为0), 待确认零申报 (同上, 仅待确认人员)。
    """
    import os
    from datetime import datetime as _dt

    os.makedirs(output_dir, exist_ok=True)
    timestamp = _dt.now().strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(output_dir, f"增减员比对结果_{pay_month}_{timestamp}.xlsx")

    wb = Workbook()
    ws_add = wb.active
    ws_add.title = "增员名单"
    ws_add.append(COMPARE_HEADERS)
    for row in add_rows:
        ws_add.append(row)

    ws_departed = wb.create_sheet("近期离职人员")
    ws_departed.append(COMPARE_HEADERS)
    for row in departed_rows:
        ws_departed.append(row)

    ws_pending = wb.create_sheet("待确认近期离职人员")
    ws_pending.append(COMPARE_HEADERS)
    for row in pending_rows:
        ws_pending.append(row)

    if zero_rows is not None:
        ws_zero = wb.create_sheet("零申报")
        ws_zero.append(ZERO_HEADERS)
        for row in zero_rows:
            ws_zero.append(row)

    if zero_pending_rows is not None:
        ws_zero_pending = wb.create_sheet("待确认零申报")
        ws_zero_pending.append(ZERO_HEADERS)
        for row in zero_pending_rows:
            ws_zero_pending.append(row)

    if verify_rows is not None:
        ws_verify = wb.create_sheet("增员验证")
        ws_verify.append(COMPARE_HEADERS + VERIFY_HEADERS)
        for row in verify_rows:
            ws_verify.append(row)

    if remove_verify_rows is not None:
        ws_remove_verify = wb.create_sheet("减员验证")
        ws_remove_verify.append(COMPARE_HEADERS + REMOVE_VERIFY_HEADERS)
        for row in remove_verify_rows:
            ws_remove_verify.append(row)

    if tc93_rows:
        ws_tc93 = wb.create_sheet("TC93工资明细")
        ws_tc93.append(TC93_HEADERS)
        for row in tc93_rows:
            ws_tc93.append(row)

    if tc8m_rows:
        ws_tc8m = wb.create_sheet("TC8M发放明细")
        ws_tc8m.append(TC8M_HEADERS)
        for row in tc8m_rows:
            ws_tc8m.append(row)

    if tc90_rows:
        ws_tc90 = wb.create_sheet("TC90合同明细")
        ws_tc90.append(TC90_HEADERS)
        for row in tc90_rows:
            ws_tc90.append(row)

    add_explanation_sheet(wb, [
        ("增员名单", [
            "个税端本应增员但未增员的人员 = 当月发薪(TC93) ∪ 未发薪 ∪ 有合同(TC90) − 个税端导出人员 − 排除项(特殊结算单元)。",
            "以 51 列人员信息导入格式输出。",
        ]),
        ("近期离职人员", [
            "个税端在职(无合同终止日期) 且不在发薪/未发薪/有合同中，且有工资结束年月(ATC90AV) 的人员。",
            "以 51 列人员信息导入格式输出。",
        ]),
        ("待确认近期离职人员", [
            "判定同「近期离职人员」，但无工资结束年月，需人工确认是否离职。",
            "延期发放人员（申报月无发放、次月月初已发，可确认在职）除外，已归入零申报。",
        ]),
        ("零申报", [
            "个税端在职且无发薪、未判定为离职/待确认的人员，按正常工资薪金所得 29 列格式输出，本期收入及养老/医疗/失业/公积金填 0，其余列留空供零申报。",
            "含延期发放人员：申报月已做工资但未发放、次月月初(1~10日)有发放记录，属在职待申报。",
        ]),
        ("待确认零申报", [
            "与「零申报」相同，仅针对待确认离职人员，供核实后零申报。",
        ]),
        ("增员验证", [
            "在 51 列基础上追加验证列（TC93 工资四要素、状态、对比来源、增减员标志等），回溯增员判定依据。",
        ]),
        ("减员验证", [
            "在 51 列基础上追加验证列，含合同开始日期（ATC90C）、合同终止日期（ATC90D）、工资结束年月（ATC90AV）等，回溯减员判定依据。",
        ]),
        ("TC93工资明细 / TC8M发放明细 / TC90合同明细", [
            "参与比对人员的工资(TC93)、发放(TC8M)、合同(TC90) 原始明细，用于核对判定口径。",
        ]),
    ])

    wb.save(output_path)
    return GenerateResult(
        file_path=output_path,
        template_type="增减员比对结果",
        record_count=len(add_rows) + len(departed_rows) + len(pending_rows),
        validation_pass=0,
        validation_fail=0,
    )