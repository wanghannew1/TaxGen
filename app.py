"""Flask 主应用 - 个税模板填表工具"""
import atexit
import os
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from db import init_db, get_connection, close_db
from queries import get_available_months, get_salary_records, get_personnel_info, get_suggestions, search_tc8m, get_abnormal_records, get_tc93_all_fields, get_tc93_field_comments, get_merge_warnings, get_pay_months, get_payroll_cert_numbers, get_tc90_termination_dates, get_payroll_personnel
from templates_gen.normal_salary import generate_normal_salary, generate_tc93_full_sheet, generate_abnormal_sheet
from templates_gen.labor_service import generate_labor_service
from templates_gen.annual_bonus import generate_annual_bonus
from templates_gen.personnel_info import generate_personnel_info
from templates_gen.validation import validate_salary_records

app = Flask(__name__)
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 注意: 连接池在进程生命周期内保持打开 (oracledb 连接在 GC 时自动归还池),
# 仅在进程退出时关闭。不能在 teardown_appcontext 中调用 close_db(),
# 否则每个请求后池被关闭, 后续所有数据库请求都会失败。
atexit.register(close_db)


def merge_records_by_person(records, by_pay_month: bool = False):
    """按人合并多条工资记录为一笔，基准取时间上最后一个批次(批次号最大、流水号最大)。

    by_pay_month=False: 按人+所属月份合并（现状，同人同月多笔合并，同一人可能多行）。
    by_pay_month=True:  按人+发放月份合并（同一发放月份内每人一行，跨所属月份的收入/五险一金/个税全部合计）。
                        仅适用于组合确认流程——所有组合共享同一发放月份(TC8M.ATC8G7)。
    """
    from copy import deepcopy
    group_key = (lambda rec: (rec.职工号,)) if by_pay_month \
        else (lambda rec: (rec.职工号, rec.工资所属年月))
    groups = {}
    for rec in records:
        groups.setdefault(group_key(rec), []).append(rec)

    merged = []
    for key, recs in groups.items():
        if by_pay_month:
            # 跨所属月份合并时"最后批次"语义失效，基准取流水号最大(最新经办)的记录
            base = max(recs, key=lambda r: r.tc930_id)
        else:
            base = max(recs, key=lambda r: (int(r.当月批次 or 0), r.tc930_id))
        m = deepcopy(base)
        for rec in recs:
            if rec is base:
                continue
            m.应发工资 += rec.应发工资
            m.实发工资 += rec.实发工资
            m.个人所得税 += rec.个人所得税
            m.工资总额 += rec.工资总额
            m.独生子女费 += rec.独生子女费
            m.采暖费 += rec.采暖费
            m.奖金 += rec.奖金
            m.养老个人 += rec.养老个人
            m.医疗个人 += rec.医疗个人
            m.失业个人 += rec.失业个人
            m.公积金个人 += rec.公积金个人
            m.补缴及退款保险金额个人 += rec.补缴及退款保险金额个人
            m.大病险个人 += rec.大病险个人
            m.补发3 += rec.补发3
            m.个人交纳现金 += rec.个人交纳现金
            m.个人其他调整 += rec.个人其他调整
            m.个人欠款 += rec.个人欠款
            m.扣款大病险 += rec.扣款大病险
            m.税后工会会费 += rec.税后工会会费
            m.个人代理费 += rec.个人代理费
            m.意外险个人 += rec.意外险个人
        merged.append(m)
    return merged

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/months")
def api_months():
    try:
        conn = get_connection()
        months = get_available_months(conn)
        return jsonify([{"value": m.value, "label": m.label} for m in months])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/suggestions/<int:month>")
def api_suggestions(month):
    try:
        conn = get_connection()
        combos = get_suggestions(conn, month)

        return jsonify({
            "month": month,
            "combos": [{
                "unit": c["unit"],
                "unit_name": c["unit_name"],
                "salary_month": c["salary_month"],
                "seq": c["seq"],
                "person_count": c["person_count"],
                "total_income": round(c["total_income"], 2)
            } for c in combos],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tc8m/search")
def api_tc8m_search():
    try:
        conn = get_connection()
        unit_name = request.args.get("unit_name", "").strip()
        salary_month = int(request.args.get("salary_month", 0) or 0)
        pay_month = int(request.args.get("pay_month", 0) or 0)
        seq = request.args.get("seq", "").strip()
        status = int(request.args.get("status", -1) or -1)
        handler = request.args.get("handler", "").strip()
        results = search_tc8m(conn, unit_name, salary_month, pay_month, seq, status, handler)
        return jsonify({"results": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/generate", methods=["POST"])
def api_generate():
    try:
        data = request.get_json()
        month = data.get("month")
        templates = data.get("templates", [])
        confirmed_combos = data.get("confirmed_combos")
        
        if not month:
            return jsonify({"error": "请选择月份"}), 400
        if not templates:
            return jsonify({"error": "请选择至少一个模板"}), 400
        
        conn = get_connection()
        if confirmed_combos is not None:
            combo_set = {(c["unit"], c["salary_month"], c["seq"]) for c in confirmed_combos}
            salary_months = {c["salary_month"] for c in confirmed_combos}
            records = []
            for sm in salary_months:
                records.extend(get_salary_records(conn, sm))
            records = [r for r in records if (r.结算单元, r.工资所属年月, r.当月批次) in combo_set]
            tc93_all = []
            abnormal = []
            for sm in salary_months:
                tc93_all.extend(r for r in get_tc93_all_fields(conn, sm)
                                if (r.get("ATB930"), r.get("ATC931"), r.get("ATC937")) in combo_set)
                abnormal.extend(r for r in get_abnormal_records(conn, sm)
                                if (r.get("ATB930"), r.get("ATC931"), r.get("ATC937")) in combo_set)
            personnel = []
            for sm in salary_months:
                personnel.extend(get_personnel_info(conn, sm))
        else:
            records = get_salary_records(conn, month)
            personnel = get_personnel_info(conn, month)
            tc93_all = get_tc93_all_fields(conn, month)
            abnormal = get_abnormal_records(conn, month)
        merge_by_person = data.get("merge_by_person", True)
        merge_by_pay_month = bool(data.get("merge_by_pay_month", True))
        raw_records = records
        if merge_by_person:
            records = merge_records_by_person(records, by_pay_month=merge_by_pay_month)
        warnings = []
        if confirmed_combos and merge_by_person:
            persons = list({r.职工号 for r in raw_records})
            warnings = get_merge_warnings(conn, [month],
                                          combo_set if confirmed_combos else set(), persons)
        abnormal_reasons = {r.get("ATC930"): f"ATC93G={r.get('ATC93G', 'NULL')}(未结算)" for r in abnormal}
        
        results = []
        for tpl in templates:
            if tpl == "normalSalary":
                if confirmed_combos:
                    top = sorted(confirmed_combos, key=lambda c: c.get("person_count", 0), reverse=True)
                    top_names = [c.get("unit_name", "") for c in top[:2] if c.get("unit_name")]
                    months_in = sorted({c["salary_month"] for c in confirmed_combos})
                    month_range = f"{months_in[0]}-{months_in[-1]}" if len(months_in) > 1 else str(months_in[0])
                    if len(confirmed_combos) == 1:
                        file_title = f"{top_names[0]}-{month_range}-{top[0].get('seq', '')}"
                    else:
                        file_title = f"{'、'.join(top_names)}等{len(confirmed_combos)}个单位{month_range}工资"
                else:
                    file_title = f"劳务派遣人员工资发放表{month}"
                r = generate_normal_salary(records, file_title, OUTPUT_DIR,
                                           tc93_all=tc93_all, abnormal=abnormal,
                                           abnormal_reasons=abnormal_reasons,
                                           combos=confirmed_combos,
                                            tc93_comments=get_tc93_field_comments(conn),
                                            raw_records=raw_records if merge_by_person else None,
                                            merge_mode="pay_month" if merge_by_pay_month else "month")
            elif tpl == "laborService":
                r = generate_labor_service(records, f"劳务派遣人员工资发放表{month}", OUTPUT_DIR)
            elif tpl == "annualBonus":
                r = generate_annual_bonus(records, f"劳务派遣人员工资发放表{month}", OUTPUT_DIR)
            elif tpl == "personnelInfo":
                r = generate_personnel_info(personnel, f"劳务派遣人员工资发放表{month}", OUTPUT_DIR)
            else:
                continue
            results.append({
                "name": os.path.basename(r.file_path),
                "type": r.template_type,
                "count": r.record_count,
                "validation_pass": r.validation_pass,
                "validation_fail": r.validation_fail,
                "download_url": f"/api/download/{os.path.basename(r.file_path)}"
            })
        
        return jsonify({
            "files": results,
            "abnormal_count": len(abnormal),
            "tc93_total_count": len(tc93_all),
            "merge_warnings": warnings
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/download/<filename>")
def api_download(filename):
    try:
        filepath = os.path.join(OUTPUT_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({"error": "文件不存在"}), 404
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/validate/<int:month>")
def api_validate(month):
    try:
        conn = get_connection()
        records = get_salary_records(conn, month)
        report = validate_salary_records(records)
        return jsonify({
            "total_count": report.total_count,
            "pass_count": report.pass_count,
            "fail_count": report.fail_count,
            "pass_rate": f"{report.pass_count/report.total_count*100:.1f}%" if report.total_count > 0 else "0%",
            "details": report.details[:50]  # 只返回前50条失败记录
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/tax-return")
def page_tax_return():
    return render_template("tax_return.html")

@app.route("/api/pay-months")
def api_pay_months():
    try:
        conn = get_connection()
        return jsonify([{"value": m.value, "label": m.label} for m in get_pay_months(conn)])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tax-return/import", methods=["POST"])
def api_tax_return_import():
    try:
        from tax_return import init_db as init_return_db, parse_return_file, import_records
        init_return_db()
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "请选择回盘文件"}), 400
        tmp_path = os.path.join(OUTPUT_DIR, "_return_tmp" + os.path.splitext(file.filename)[1])
        file.save(tmp_path)
        try:
            records = parse_return_file(tmp_path)
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        if not records:
            return jsonify({"error": "未解析到有效记录，请检查文件格式"}), 400
        months = sorted({r["month"] for r in records})
        import_records(records)
        return jsonify({"ok": True, "count": len(records), "months": months})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tax-return/status")
def api_tax_return_status():
    try:
        from tax_return import summarize
        month = int(request.args.get("month", 0) or 0)
        if not month:
            return jsonify({"error": "请选择月份"}), 400
        conn = get_connection()
        result = summarize(conn, month)
        result.pop("details", None)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tax-return/details")
def api_tax_return_details():
    try:
        from tax_return import compare_month
        month = int(request.args.get("month", 0) or 0)
        status = request.args.get("status", "")
        search = request.args.get("search", "").strip()
        page = max(1, int(request.args.get("page", 1) or 1))
        page_size = min(500, max(1, int(request.args.get("page_size", 50) or 50)))
        if not month:
            return jsonify({"error": "请选择月份"}), 400
        conn = get_connection()
        details, combo_stats = compare_month(conn, month)
        if status:
            details = [d for d in details if d["status"] == status]
        if search:
            details = [d for d in details
                       if search.lower() in d["name"].lower() or search in d["cert_no"]]
        total = len(details)
        start = (page - 1) * page_size
        return jsonify({"total": total, "page": page, "page_size": page_size,
                        "details": details[start:start + page_size]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tax-return/export")
def api_tax_return_export():
    try:
        from openpyxl import Workbook
        from tax_return import compare_month
        month = int(request.args.get("month", 0) or 0)
        if not month:
            return jsonify({"error": "请选择月份"}), 400
        conn = get_connection()
        details, combo_stats = compare_month(conn, month)
        wb = Workbook()
        ws = wb.active
        ws.title = "报税状态统计"
        ws.append(["月份", "姓名", "证件号码", "结算单元", "所属月份-批次", "系统本期收入", "回盘本期收入", "收入差异", "系统预算个税", "回盘个税(应补退)", "个税差值", "状态"])
        for d in details:
            ws.append([month, d["name"], d["cert_no"], d.get("unit_name", ""), d.get("unit_periods", ""),
                       d["sys_income"], d.get("ret_income"), d.get("income_diff"), d["sys_tax"],
                       d.get("ret_tax"), d["diff"], d["status"]])
        ws2 = wb.create_sheet("按结算单元批次")
        ws2.append(["结算单元", "结算单元名称", "所属月份", "批次", "人数", "已报送", "未报送"])
        for c in combo_stats:
            ws2.append([c["unit"], c.get("unit_name", ""), c["month"], c["seq"], c["count"], c["reported"], c["unreported"]])
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"报税状态统计_{month}_{timestamp}.xlsx"
        wb.save(os.path.join(OUTPUT_DIR, filename))
        return jsonify({"download_url": f"/api/download/{filename}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/personnel-compare")
def page_personnel_compare():
    return render_template("personnel_compare.html")


@app.route("/special-units")
def page_special_units():
    return render_template("special_units.html")


@app.route("/api/special-units", methods=["GET"])
def api_special_units_list():
    """获取特殊结算单元配置列表 (SQLite config_db)。"""
    try:
        from config_db import get_special_units
        return jsonify({"units": get_special_units()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/special-units", methods=["POST"])
def api_special_units_add():
    """新增特殊结算单元配置 (exclude_all=完全排除不增员不报税, SQLite config_db)。"""
    try:
        from config_db import add_special_unit
        data = request.get_json()
        unit_code = int(data.get("unit_code", 0) or 0)
        unit_name = str(data.get("unit_name", "") or "")
        exclude_all = bool(data.get("exclude_all", False))
        if not unit_code:
            return jsonify({"error": "请填写结算单元代码"}), 400
        add_special_unit(unit_code, unit_name, exclude_all=exclude_all)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/special-units/<int:unit_code>/mode", methods=["POST"])
def api_special_units_mode(unit_code):
    """更新特殊结算单元配置的排除模式 (zero_salary_no_add / exclude_all 开关, SQLite)。"""
    try:
        from config_db import update_special_unit
        data = request.get_json() or {}
        exclude_all = data.get("exclude_all")
        zero_salary_no_add = data.get("zero_salary_no_add")
        update_special_unit(unit_code,
                            exclude_all=exclude_all if exclude_all is not None else None,
                            zero_salary_no_add=zero_salary_no_add if zero_salary_no_add is not None else None)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/special-units/template")
def api_special_units_template():
    """下载特殊结算单元配置导入模板 (4列: 代码/名称/工资为0/完全排除)。"""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "特殊结算单元配置"
        ws.append(["结算单元代码", "结算单元名称", "工资为0不增员不报税", "完全排除不增员不报税"])
        ws.append([None, "测试A-仅工资0", 1, 0])
        ws.append([None, "测试B-仅完全排除", 0, 1])
        ws.append([None, "测试C-两种", 1, 1])
        from datetime import datetime as _dt
        filename = f"特殊结算单元配置导入模板_{_dt.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(os.path.join(OUTPUT_DIR, filename))
        return jsonify({"download_url": f"/api/download/{filename}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/special-units/unit-list-export")
def api_special_units_unit_list_export():
    """导出结算单元代码-名称对照表 (供填写导入配置用)。"""
    try:
        from openpyxl import Workbook
        from queries import get_units
        conn = get_connection()
        units = get_units(conn)  # 全部结算单元
        wb = Workbook()
        ws = wb.active
        ws.title = "结算单元对照表"
        ws.append(["结算单元代码", "结算单元名称"])
        for u in units:
            ws.append([u["code"], u["name"]])
        from datetime import datetime as _dt
        filename = f"结算单元对照表_{_dt.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(os.path.join(OUTPUT_DIR, filename))
        return jsonify({"download_url": f"/api/download/{filename}", "count": len(units)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/special-units/export")
def api_special_units_export():
    """导出特殊结算单元配置为 Excel (SQLite config_db)。"""
    try:
        from openpyxl import Workbook
        from config_db import get_special_units
        units = get_special_units()
        wb = Workbook()
        ws = wb.active
        ws.title = "特殊结算单元配置"
        ws.append(["结算单元代码", "结算单元名称", "工资为0不增员不报税", "完全排除不增员不报税"])
        for u in units:
            ws.append([u["code"], u["name"], u["zero_salary_no_add"], u["exclude_all"]])
        from datetime import datetime as _dt
        filename = f"特殊结算单元配置_{_dt.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(os.path.join(OUTPUT_DIR, filename))
        return jsonify({"download_url": f"/api/download/{filename}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/special-units/import", methods=["POST"])
def api_special_units_import():
    """从 Excel 导入特殊结算单元配置 (覆盖式)。

    结算单元代码可留空, 按结算单元名称自动匹配; 名称匹配不到或多个时跳过并提示。
    """
    try:
        from openpyxl import load_workbook
        from config_db import delete_special_unit, add_special_unit_full
        from queries import lookup_unit_codes_by_name
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "请上传配置文件"}), 400
        tmp_path = os.path.join(OUTPUT_DIR, "_special_units_tmp.xlsx")
        file.save(tmp_path)
        conn = get_connection()
        try:
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active
            units = []
            skipped = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                # 兼容两种格式:
                # 4列: 结算单元代码 | 结算单元名称 | 工资为0不增员不报税 | 完全排除不增员不报税
                # 3列: 配置(名称或代码) | 工资为0不增员 | 完全排除不增员
                vals = list(row)
                while len(vals) < 4:
                    vals.append(None)
                if vals[2] is not None or vals[3] is not None:
                    code_raw = vals[0]
                    name = str(vals[1] or "")
                    zero_flag = int(vals[2] or 0)
                    exclude_all = int(vals[3] or 0)
                    code = int(code_raw) if str(code_raw or "").strip().isdigit() else None
                else:
                    code_raw = vals[0]
                    name = str(code_raw or "")
                    zero_flag = int(vals[1] or 0)
                    exclude_all = int(vals[2] or 0)
                    code = int(code_raw) if str(code_raw or "").strip().isdigit() else None
                if not code:
                    # 代码留空时按名称自动匹配
                    codes = lookup_unit_codes_by_name(conn, name)
                    if len(codes) == 1:
                        code = codes[0]
                    else:
                        skipped.append(f"{name}(匹配到{len(codes)}个代码)")
                        continue
                units.append({"code": code, "name": name,
                              "zero_salary_no_add": zero_flag, "exclude_all": exclude_all})
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        delete_special_unit(None)  # 清空 (SQLite 全部删除)
        for u in units:
            add_special_unit_full(u["code"], u["name"],
                                  zero_salary_no_add=u["zero_salary_no_add"],
                                  exclude_all=u["exclude_all"])
        resp = {"ok": True, "count": len(units)}
        if skipped:
            resp["skipped"] = skipped
        return jsonify(resp)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/special-units/<int:unit_code>", methods=["DELETE"])
def api_special_units_delete(unit_code):
    """删除特殊结算单元配置 (SQLite)。"""
    try:
        from config_db import delete_special_unit
        delete_special_unit(unit_code)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/payroll-records")
def api_payroll_records():
    """查询工资发放情况 (TC8M, 按经办年月范围, 已发放)。"""
    try:
        from queries import get_payroll_records
        start = int(request.args.get("start", 0) or 0)
        end = int(request.args.get("end", 0) or 0)
        if not start or not end:
            return jsonify({"error": "请选择发薪月份(起始)和(结束)"}), 400
        if end < start:
            return jsonify({"error": "发薪月份(结束)不能早于(起始)"}), 400
        conn = get_connection()
        records = get_payroll_records(conn, start, end)
        return jsonify({"records": records, "count": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/personnel-compare/filters")
def api_personnel_compare_filters():
    """查询经办人/结算单元/单位列表 (供筛选下拉框)。"""
    try:
        from queries import get_handlers, get_units, get_depts
        conn = get_connection()
        pay_month = int(request.args.get("pay_month", 0) or 0)
        return jsonify({
            "handlers": get_handlers(conn, pay_month),
            "units": get_units(conn, pay_month),
            "depts": get_depts(conn, pay_month),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/personnel-compare/default-report-month")
def api_personnel_compare_default_month():
    """查询默认上报发薪月份 (1-15日上月, 16-月末本月)。"""
    try:
        from queries import get_default_report_month
        conn = get_connection()
        return jsonify({"month": get_default_report_month(conn)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/personnel-compare/latest-pay-date")
def api_personnel_compare_latest_pay_date():
    """查询最近一笔工资发放日期 (TC8M.AAE036)。"""
    try:
        from queries import get_latest_pay_date
        conn = get_connection()
        pay_month, pay_date = get_latest_pay_date(conn)
        return jsonify({
            "pay_month": pay_month,
            "pay_date": pay_date.strftime("%Y-%m-%d") if pay_date else None,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/personnel-compare/compare", methods=["POST"])
def api_personnel_compare():
    """上传个税端导出文件 + 选择最近1~2次发薪月份 → 增减员比对 → 生成 Excel。

    离职时间截止日期 (termination_deadline, YYYY-MM-DD) 用于过滤 TC90 离职日期:
    超过截止日期的合同终止日期视为未到期, 归入待确认。

    参数:
    - file: 个税端导出文件 (.xls)
    - pay_month_start/pay_month_end: 发薪月份时间段 (显式起止)
    - unpaid_month_start/unpaid_month_end: 未发薪工资表所属月份时间段 (显式起止)
    - pay_start_time/pay_end_time: 可选发薪经办时间 (精确到时分秒, 默认关闭)
    - contract_start_time/contract_end_time: 合同签署时间范围 (精确到时分秒, 可选)
    - termination_deadline: 离职时间截止日期 (默认最近发薪日期)
    """
    try:
        from queries import (get_payroll_cert_numbers, get_tc90_termination_dates,
                             get_payroll_personnel, get_unpaid_salary_persons,
                             get_pay_month_range, get_unpaid_month_range,
                             get_contract_signed_persons, get_contract_date_range,
                             get_latest_pay_date, get_personnel_by_certs)
        from templates_gen.personnel_compare import compare_personnel, generate_compare_excel
        from templates_gen.tax_export_parser import parse_tax_export

        file = request.files.get("file")
        if not file:
            return jsonify({"error": "请上传个税端导出文件"}), 400
        pay_month_start = int(request.form.get("pay_month_start", 0) or 0)
        if not pay_month_start:
            return jsonify({"error": "请选择发薪月份(起始)"}), 400
        pay_month_end = int(request.form.get("pay_month_end", 0) or pay_month_start)
        if pay_month_end < pay_month_start:
            return jsonify({"error": "发薪月份(结束)不能早于(起始)"}), 400
        unpaid_month_start = int(request.form.get("unpaid_month_start", 0) or 0)
        unpaid_month_end = int(request.form.get("unpaid_month_end", 0) or unpaid_month_start)
        if unpaid_month_start and unpaid_month_end < unpaid_month_start:
            return jsonify({"error": "未发薪所属月份(结束)不能早于(起始)"}), 400
        contract_start_time = request.form.get("contract_start_time", "").strip()
        contract_end_time = request.form.get("contract_end_time", "").strip()
        from datetime import datetime as _dt
        contract_start_dt = _dt.strptime(contract_start_time, "%Y-%m-%dT%H:%M") if contract_start_time else None
        contract_end_dt = _dt.strptime(contract_end_time, "%Y-%m-%dT%H:%M") if contract_end_time else None
        # 可选经办时间过滤 (精确到时分秒, 默认关闭)
        pay_start_time = request.form.get("pay_start_time", "").strip()
        pay_end_time = request.form.get("pay_end_time", "").strip()
        from datetime import datetime as _dt
        pay_start_dt = _dt.strptime(pay_start_time, "%Y-%m-%dT%H:%M") if pay_start_time else None
        pay_end_dt = _dt.strptime(pay_end_time, "%Y-%m-%dT%H:%M") if pay_end_time else None
        filter_handlers = [h for h in request.form.getlist("handler") if h.strip()]
        filter_units = [int(u) for u in request.form.getlist("unit") if u.strip()]
        filter_depts = [d for d in request.form.getlist("dept") if d.strip()]
        deadline = request.form.get("termination_deadline", "").strip()
        deadline_date = None
        if deadline:
            try:
                deadline_date = _dt.strptime(deadline, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "离职时间截止日期格式错误 (应为 YYYY-MM-DD)"}), 400
        ext = os.path.splitext(file.filename or "")[1].lower()
        if ext != ".xls":
            return jsonify({"error": "仅支持 .xls 格式的个税端导出文件"}), 400

        tmp_path = os.path.join(OUTPUT_DIR, "_compare_tmp" + ext)
        file.save(tmp_path)
        try:
            tax_export_persons = parse_tax_export(tmp_path)
        except Exception as e:
            return jsonify({"error": f"解析个税端导出文件失败: {e}"}), 400
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        if not tax_export_persons:
            return jsonify({"error": "导出文件中未解析到有效人员记录"}), 400

        conn = get_connection()
        # 发薪月份范围: 用户显式选择 [起始, 结束]
        pay_months = list(range(pay_month_start, pay_month_end + 1))
        payroll_certs = set()
        payroll_personnel = []
        seen = set()
        for m in pay_months:
            payroll_certs |= get_payroll_cert_numbers(conn, m, start_time=pay_start_dt, end_time=pay_end_dt)
            for p in get_payroll_personnel(conn, m, start_time=pay_start_dt, end_time=pay_end_dt):
                if p.身份证 and p.身份证 not in seen:
                    seen.add(p.身份证)
                    payroll_personnel.append(p)
        # 未发薪工资表: 用户显式选择 [起始, 结束]
        unpaid_persons = set()
        unpaid_months = []
        if unpaid_month_start:
            unpaid_months = list(range(unpaid_month_start, unpaid_month_end + 1))
            unpaid_persons = get_unpaid_salary_persons(conn, unpaid_months)
        # 合同签署: 合同开始日期在起始~结束时间范围内 (精确到时分秒)
        contract_persons = set()
        if contract_start_dt or contract_end_dt:
            contract_persons = get_contract_signed_persons(conn, contract_start_dt, contract_end_dt)
        # 疑似近期离职: 个税端未标记离职(无离职日期)且不在发薪/未发薪/合同名单
        active_certs = {
            str(p.get("证件号码") or "").strip().upper()
            for p in tax_export_persons
            if not str(p.get("离职日期") or "").strip()
        }
        protected = payroll_certs | unpaid_persons | contract_persons
        suspect_certs = active_certs - protected
        termination_dates = get_tc90_termination_dates(conn, suspect_certs)
        # 离职日期超过截止日期的视为合同未到期, 不列为已离职
        if deadline_date:
            termination_dates = {
                c: d for c, d in termination_dates.items()
                if d.date() <= deadline_date
            }
        add_rows, departed_rows, pending_rows, stats = compare_personnel(
            tax_export_persons, payroll_certs, payroll_personnel, termination_dates,
            unpaid_persons=unpaid_persons, contract_signed_persons=contract_persons)
        # 特殊结算单元: 工资为0不增员不报税 + 完全排除不增员不报税
        # 配置存 SQLite (config_db), Oracle 只读
        from config_db import get_zero_salary_unit_codes, get_excluded_unit_codes
        from queries import get_zero_salary_certs, get_excluded_unit_certs, get_person_units
        exclude_certs = set()
        zero_codes = get_zero_salary_unit_codes()
        excl_codes = get_excluded_unit_codes()
        if zero_codes:
            exclude_certs |= get_zero_salary_certs(conn, pay_months, unpaid_months, zero_codes)
        if excl_codes:
            relevant_months = sorted(set(pay_months) | set(unpaid_months))
            exclude_certs |= get_excluded_unit_certs(conn, pay_months, excl_codes,
                                                     relevant_months=relevant_months)
        # 经办人/结算单元/单位过滤 + 备注(结算单元名称)数据
        tax_cert_set = {str(p.get("证件号码") or "").strip().upper()
                        for p in tax_export_persons}
        add_candidates = (payroll_certs | unpaid_persons | contract_persons) - tax_cert_set
        # 过滤时减员候选也需要人员归属信息 (个税端未标记离职人员)
        unit_query_certs = set(add_candidates)
        if filter_handlers or filter_units or filter_depts:
            active_certs = {str(p.get("证件号码") or "").strip().upper()
                            for p in tax_export_persons
                            if not str(p.get("离职日期") or "").strip()}
            unit_query_certs |= active_certs
        person_units = get_person_units(conn, unit_query_certs, pay_months)
        # 仅合同人员(无工资记录)补充: TC90 结算单元, 降级单位名称
        missing_certs = {c for c in unit_query_certs
                         if not person_units.get(c, {}).get("unit_name")}
        if missing_certs:
            from queries import get_person_units_contract
            person_units.update(get_person_units_contract(conn, missing_certs))
        if exclude_certs or filter_handlers or filter_units or filter_depts:
            add_rows, departed_rows, pending_rows, stats = compare_personnel(
                tax_export_persons, payroll_certs, payroll_personnel, termination_dates,
                unpaid_persons=unpaid_persons, contract_signed_persons=contract_persons,
                person_units=person_units, filter_handlers=filter_handlers,
                filter_units=filter_units, filter_depts=filter_depts,
                exclude_certs=exclude_certs)
        # 补充增员人员详细信息 (未发薪/合同签署人员不在 payroll_personnel 中)
        if stats["add_count"] > len(add_rows):
            from templates_gen.personnel_compare import IDX_证件号码, map_personnel_info_to_row
            have_certs = {r[IDX_证件号码] for r in add_rows}
            need_certs = stats["add_certs"] - have_certs
            if need_certs:
                extra_people = get_personnel_by_certs(conn, need_certs)
                for p in extra_people:
                    row = map_personnel_info_to_row(p)
                    info = (person_units or {}).get(str(p.身份证 or "").strip().upper())
                    if info:
                        # 备注优先结算单元名称(ATB931); 降级单位名称时标明"非结算单元名称"
                        remark = info.get("unit_name") or ""
                        if remark:
                            row[25] = remark
                        elif info.get("dept_name"):
                            row[25] = f"单位名称（非结算单元名称）：{info['dept_name']}"
                    add_rows.append(row)
        # 增员名单/验证按备注(结算单元)排序, 相同结算单元相邻
        add_rows.sort(key=lambda r: str(r[25] or ""))
        # 增员验证 Sheet: 为每个增员人员组装验证行 + TC93/TC8M/TC90 明细
        from queries import get_salary_details, get_tc8m_records, get_tc90_records
        from templates_gen.personnel_compare import (IDX_证件号码, build_verify_row,
                                                     map_personnel_info_to_row as _map_row)
        add_certs_final = {r[IDX_证件号码] for r in add_rows}
        verify_params = {
            "pay_months": pay_months,
            "unpaid_months": unpaid_months,
            "contract_start": contract_start_dt.strftime("%Y-%m-%d %H:%M") if contract_start_dt else "",
            "contract_end": contract_end_dt.strftime("%Y-%m-%d %H:%M") if contract_end_dt else "",
        }
        salary_details = get_salary_details(conn, add_certs_final,
                                            sorted(set(pay_months + unpaid_months)))
        tc8m_details = get_tc8m_records(conn, add_certs_final, pay_months)
        tc90_details = get_tc90_records(conn, add_certs_final)
        # 发薪工资明细: 按 TC8M 发薪记录的所属月份范围查询 (可能早于发薪月份)
        paid_salary_months = sorted({r["salary_month"] for r in tc8m_details
                                     if r.get("salary_month")})
        paid_salary_details = get_salary_details(conn, add_certs_final, paid_salary_months) if paid_salary_months else []
        # 合同开始日期: 取 TC90 最早的 ATC90C
        contract_start_map = {}
        for r in tc90_details:
            cert = r["cert"]
            if cert not in contract_start_map or (r["合同开始日期"] and r["合同开始日期"] < contract_start_map[cert]):
                contract_start_map[cert] = r["合同开始日期"]
        salary_by_cert = {}
        for r in salary_details:
            salary_by_cert.setdefault(r["cert"], []).append(r)
        paid_salary_by_cert = {}
        for r in paid_salary_details:
            paid_salary_by_cert.setdefault(r["cert"], []).append(r)
        tc8m_by_cert = {}
        for r in tc8m_details:
            tc8m_by_cert.setdefault(r["cert"], []).append(r)
        tc90_by_cert = {}
        for r in tc90_details:
            tc90_by_cert.setdefault(r["cert"], []).append(r)
        verify_rows = []
        for r in add_rows:
            cert = r[IDX_证件号码]
            verify_rows.append(build_verify_row(
                r, cert, verify_params,
                paid_salary_by_cert.get(cert, []),
                salary_by_cert.get(cert, []),
                tc8m_by_cert.get(cert, []),
                contract_start_map.get(cert),
                contract_details=tc90_by_cert.get(cert, [])))
        # 减员验证 Sheet: 为近期离职/待确认离职人员组装验证行
        from templates_gen.personnel_compare import build_remove_verify_row
        remove_all = [(r, "近期离职") for r in departed_rows] + [(r, "待确认近期离职") for r in pending_rows]
        remove_certs = {r[IDX_证件号码] for r, _ in remove_all}
        remove_verify_rows = []
        if remove_certs:
            remove_salary = get_salary_details(conn, remove_certs,
                                               sorted(set(pay_months + unpaid_months)))
            remove_tc8m = get_tc8m_records(conn, remove_certs, pay_months)
            remove_tc90 = get_tc90_records(conn, remove_certs)
            remove_salary_by_cert = {}
            for r in remove_salary:
                remove_salary_by_cert.setdefault(r["cert"], []).append(r)
            remove_tc8m_by_cert = {}
            for r in remove_tc8m:
                remove_tc8m_by_cert.setdefault(r["cert"], []).append(r)
            remove_tc90_by_cert = {}
            for r in remove_tc90:
                remove_tc90_by_cert.setdefault(r["cert"], []).append(r)
            remove_contract_start = {}
            remove_contract_end = {}
            for r in remove_tc90:
                c = r["cert"]
                if c not in remove_contract_start or (r["合同开始日期"] and r["合同开始日期"] < remove_contract_start[c]):
                    remove_contract_start[c] = r["合同开始日期"]
            # 离职日期: 用已按截止日期过滤的 termination_dates (判定口径)
            # 待确认人员 (无有效终止日期) 离职日期为空
            for c, dt in termination_dates.items():
                if c in remove_contract_end:
                    remove_contract_end[c] = max(remove_contract_end[c], dt)
                else:
                    remove_contract_end[c] = dt
            for row, rtype in remove_all:
                cert = row[IDX_证件号码]
                remove_verify_rows.append(build_remove_verify_row(
                    row, cert, rtype, verify_params,
                    [],  # 减员无发薪记录
                    remove_salary_by_cert.get(cert, []),
                    remove_tc8m_by_cert.get(cert, []),
                    remove_contract_start.get(cert),
                    remove_contract_end.get(cert),
                    contract_details=remove_tc90_by_cert.get(cert, [])))
        tc93_all = []
        seen_tc93 = set()
        for r in paid_salary_details + salary_details:
            key = (r["cert"], r["salary_month"], r["seq"])
            if key in seen_tc93:
                continue
            seen_tc93.add(key)
            tc93_all.append(r)
        tc93_rows = [[r["cert"], r["姓名"], r["unit_code"], r["unit_name"], r["salary_month"], r["seq"],
                      r["应发工资"], r["本期收入"], r["养老"], r["医疗"], r["失业"], r["公积金"]]
                     for r in tc93_all]
        tc8m_all = []
        seen_tc8m = set()
        for r in tc8m_details:
            key = (r["cert"], r["salary_month"], r["seq"])
            if key in seen_tc8m:
                continue
            seen_tc8m.add(key)
            tc8m_all.append(r)
        tc8m_rows = [[r["cert"], r["姓名"], r["unit_code"], r["unit_name"], r["pay_month"],
                      r["salary_month"], r["seq"], r["handler"]] for r in tc8m_all]
        # TC90 每人最多 2 条 (按合同开始日期倒序取最近)
        tc90_by_cert = {}
        for r in tc90_details:
            tc90_by_cert.setdefault(r["cert"], []).append(r)
        tc90_rows = []
        for cert, recs in tc90_by_cert.items():
            recs_sorted = sorted(recs, key=lambda x: x["合同开始日期"], reverse=True)
            for r in recs_sorted[:2]:
                tc90_rows.append([r["cert"], r["姓名"], r["unit_code"], r["unit_name"],
                                  r["合同开始日期"], r["合同终止日期"], r["单位名称"], r["经办人"]])
        month_label = f"{pay_month_start}-{pay_month_end}" if pay_month_end != pay_month_start else str(pay_month_start)
        result = generate_compare_excel(add_rows, departed_rows, pending_rows, stats, OUTPUT_DIR, month_label,
                                        verify_rows=verify_rows, tc93_rows=tc93_rows,
                                        tc8m_rows=tc8m_rows, tc90_rows=tc90_rows,
                                        remove_verify_rows=remove_verify_rows)
        return jsonify({
            "add_count": stats["add_count"],
            "departed_count": stats["departed_count"],
            "pending_count": stats["pending_count"],
            "tax_total": stats["tax_total"],
            "payroll_total": stats["payroll_total"],
            "file_name": os.path.basename(result.file_path),
            "download_url": f"/api/download/{os.path.basename(result.file_path)}",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


init_db()

# 初始化自建 SQLite 配置库 (Oracle 只读, 配置数据不入 Oracle)
import config_db as _config_db
_config_db.init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=os.getenv("FLASK_DEBUG", "0") == "1")